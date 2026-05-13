#!/usr/bin/env python3
"""
脱敏脚本：读取原始 GMV 全量数据 Excel，输出脱敏 Excel + mapping.json

用法：
    python scripts/desensitize.py <原始Excel> [--output <输出目录>]

功能：
    - 项目组名替换（固定映射表）
    - 金额保持原值不脱敏
    - Sheet名和列名保持不变
    - 输出 mapping.json + desensitized_data.xlsx
"""

import argparse
import json
import os
import sys
from datetime import datetime

import openpyxl


# ========== 固定映射表（项目组名 + 业务线名） ==========
NAME_MAPPING = {
    # 项目组名称
    "全国大班": "项目组A",
    "云领学": "项目组B",
    "小图灵": "项目组C",
    "KP": "项目组D",
    "研习所": "项目组E",
    "纵横工作室": "项目组F",
    "线下店": "项目组G",
    "Deepthink": "项目组H",
    "博闻": "项目组I",
    "思维": "项目组J",
    "新世界": "项目组K",
    "强基业务": "项目组L",
    # 业务线名称
    "升学": "业务线A",
    "素养": "业务线B",
}


def _replace_names(value, mapping):
    """替换字符串中的项目组名和业务线名"""
    if not isinstance(value, str):
        return value
    result = value
    # 先替换长的名称，避免部分匹配
    for original, replacement in sorted(mapping.items(), key=lambda x: -len(x[0])):
        result = result.replace(original, replacement)
    return result


def desensitize_excel(input_path, output_dir=None):
    """
    脱敏Excel文件（仅替换项目组名，金额保持原值）

    Args:
        input_path: 原始Excel文件路径
        output_dir: 输出目录（默认与输入文件同目录）

    Returns:
        (desensitized_path, mapping_path) 输出文件路径元组
    """
    if not os.path.exists(input_path):
        print(f"❌ 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(input_path))
    os.makedirs(output_dir, exist_ok=True)

    # 用 data_only=True 读取（公式→值）
    wb = openpyxl.load_workbook(input_path, data_only=True)

    name_replaced = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row is None or max_col is None or max_row < 1:
            continue

        # 遍历所有单元格，只替换字符串
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                value = cell.value

                if value is None:
                    continue

                # 字符串处理：替换项目组名
                if isinstance(value, str):
                    new_value = _replace_names(value, NAME_MAPPING)
                    if new_value != value:
                        cell.value = new_value
                        name_replaced += 1

    # 保存脱敏Excel（保留原始文件名格式，确保read_excel能识别）
    original_basename = os.path.basename(input_path)
    desensitized_filename = f"desensitized_{original_basename}"
    desensitized_path = os.path.join(output_dir, desensitized_filename)
    wb.save(desensitized_path)

    # 生成 mapping.json
    mapping = {
        "name_mapping": NAME_MAPPING,
        "scale_factor": 1.0,  # 金额不缩放
        "original_file": original_basename,
        "desensitized_file": desensitized_filename,
        "created_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "note": "项目组名称和业务线名称脱敏，金额保持原值",
    }
    mapping_path = os.path.join(output_dir, "mapping.json")
    with open(mapping_path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)

    print(f"✅ 脱敏完成")
    print(f"  📄 脱敏数据: {desensitized_path}")
    print(f"  🔑 映射表: {mapping_path}")
    print(f"  🏷️  名称替换: {name_replaced} 处")
    print(f"  💰 金额: 保持原值（未缩放）")

    return desensitized_path, mapping_path


def main():
    parser = argparse.ArgumentParser(description="脱敏 GMV 全量数据 Excel（仅项目组名，不改金额）")
    parser.add_argument("input", help="原始Excel文件路径")
    parser.add_argument("--output", default=None, help="输出目录 (默认与输入文件同目录)")
    args = parser.parse_args()

    desensitize_excel(args.input, args.output)


if __name__ == "__main__":
    main()
