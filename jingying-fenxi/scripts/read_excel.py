#!/usr/bin/env python3
"""
经营周报数据读取脚本

从工作目录中读取最新的 GMV全量数据 Excel 文件，
提取关键 Sheet 数据并输出 JSON 格式供分析使用。

用法:
    python read_excel.py <工作目录路径> [--sheet <Sheet名>]

示例:
    python read_excel.py "D:\\02-work-LifeOS\\数据\\10-工作文件\\5-经营周报\\01-经营周报项目\\新建文件夹"
    python read_excel.py . --sheet "周营收-财务口径"
"""

import sys
import io
import os
import json
import glob
import re
import yaml
from datetime import datetime, date
import calendar

# Windows 编码处理
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("❌ 错误: 需要安装 openpyxl", file=sys.stderr)
    print("请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def find_latest_gmv_file(work_dir):
    """查找工作目录下最新的 GMV全量数据 文件（含脱敏后的文件）。

    按文件名中的日期分组，同一日期优先选脱敏版本。
    latest = 最新日期的文件，last_week = 次新日期的文件。
    """
    pattern = os.path.join(work_dir, "GMV全量数据_*.xlsx")
    raw_files = glob.glob(pattern)

    desen_patterns = [
        os.path.join(work_dir, "desensitized_data.xlsx"),
        os.path.join(work_dir, "desensitized_GMV全量数据_*.xlsx"),
    ]
    desen_files = []
    for dp in desen_patterns:
        desen_files.extend(glob.glob(dp))

    all_files = raw_files + desen_files
    if not all_files:
        return None, None, None

    # 按日期分组：date -> {"raw": path, "desen": path}
    date_groups = {}
    for fp in all_files:
        m = re.search(r'(\d{4}-\d{2}-\d{2})\.xlsx$', fp)
        if not m:
            continue
        d = m.group(1)
        if d not in date_groups:
            date_groups[d] = {"raw": None, "desen": None}
        basename = os.path.basename(fp)
        if basename.startswith("desensitized"):
            date_groups[d]["desen"] = fp
        else:
            date_groups[d]["raw"] = fp

    if not date_groups:
        return None, None, None

    # 按日期降序排列
    sorted_dates = sorted(date_groups.keys(), reverse=True)

    # 最新日期：优先选脱敏版本
    latest_date = sorted_dates[0]
    grp = date_groups[latest_date]
    latest = grp["desen"] or grp["raw"]

    # 次新日期：作为上周文件（优先脱敏，通常只有raw）
    last_week_file = None
    if len(sorted_dates) >= 2:
        lw_date = sorted_dates[1]
        lw_grp = date_groups[lw_date]
        last_week_file = lw_grp["desen"] or lw_grp["raw"]

    return latest, latest_date, last_week_file


def get_current_month_sheet_name(cutoff_date_str):
    """根据截止日期判断当月 Sheet 名"""
    if not cutoff_date_str:
        return None
    dt = datetime.strptime(cutoff_date_str, "%Y-%m-%d")
    month_names = {
        1: "1月", 2: "2月", 3: "3月", 4: "4月",
        5: "5月", 6: "6月", 7: "7月", 8: "8月",
        9: "9月", 10: "10月", 11: "11月", 12: "12月"
    }
    return f"{month_names[dt.month]}-B标"


def get_current_month_yoy_sheet_name(cutoff_date_str):
    """根据截止日期判断当月同比 Sheet 名"""
    if not cutoff_date_str:
        return None
    dt = datetime.strptime(cutoff_date_str, "%Y-%m-%d")
    month_names = {
        1: "1月", 2: "2月", 3: "3月", 4: "4月",
        5: "5月", 6: "6月", 7: "7月", 8: "8月",
        9: "9月", 10: "10月", 11: "11月", 12: "12月"
    }
    return f"{month_names[dt.month]}同比"


def load_config():
    """加载 config.yaml 阈值参数"""
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "config.yaml")
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)
    return {}


def calc_time_progress(cutoff_date_str):
    """计算时间进度基准（动态检测季度）"""
    if not cutoff_date_str:
        return {}
    dt = datetime.strptime(cutoff_date_str, "%Y-%m-%d")

    # 动态检测当前季度（Q1=1-3月，Q2=4-6月，Q3=7-9月，Q4=10-12月）
    month = dt.month
    if month <= 3:
        quarter = 1
        q_start = datetime(dt.year, 1, 1)
        q_end = datetime(dt.year, 3, 31)
    elif month <= 6:
        quarter = 2
        q_start = datetime(dt.year, 4, 1)
        q_end = datetime(dt.year, 6, 30)
    elif month <= 9:
        quarter = 3
        q_start = datetime(dt.year, 7, 1)
        q_end = datetime(dt.year, 9, 30)
    else:
        quarter = 4
        q_start = datetime(dt.year, 10, 1)
        q_end = datetime(dt.year, 12, 31)

    q_total_days = (q_end - q_start).days + 1  # 包含首尾两天
    q_elapsed = (dt - q_start).days + 1  # 截止日当天算进去
    q_progress = min(q_elapsed / q_total_days, 1.0)

    # 当月进度: 截止日当天算进去
    days_in_month = calendar.monthrange(dt.year, dt.month)[1]
    month_progress = dt.day / days_in_month

    return {
        "quarter": quarter,
        "q_progress": round(q_progress * 100, 1),
        "q_progress_display": f"{round(q_progress * 100, 1)}%",
        "month_progress": round(month_progress * 100, 1),
        "month_progress_display": f"{round(month_progress * 100, 1)}%",
        "cutoff_date": cutoff_date_str,
        "q_total_days": q_total_days,
        "q_elapsed_days": q_elapsed,
        "month_total_days": days_in_month,
        "month_elapsed_days": dt.day,
    }


def deduplicate_headers(raw_headers):
    """处理重复列名：给重复的列名加上前一个非重复列名作为前缀
    例如 ['总营收','占比','环比','拉新GMV','环比','续报GMV','环比','SMROI','环比']
    变成 ['总营收','占比','总营收_环比','拉新GMV','拉新GMV_环比','续报GMV','续报GMV_环比','SMROI','SMROI_环比']
    """
    result = []
    seen = {}
    last_unique = ""
    for idx, h in enumerate(raw_headers):
        h_str = str(h).strip() if h else ""
        if not h_str:
            # Empty header - assign a positional name
            result.append(f"_col_{idx}")
            continue
        if h_str in seen:
            # This is a duplicate - prefix it with the last unique column name
            if last_unique:
                new_name = f"{last_unique}_{h_str}"
            else:
                new_name = f"{h_str}_{seen[h_str]}"
            # Ensure uniqueness with a counter
            counter = 2
            base_name = new_name
            while new_name in seen:
                new_name = f"{base_name}_{counter}"
                counter += 1
            result.append(new_name)
            seen[new_name] = 1
            seen[h_str] += 1
        else:
            result.append(h_str)
            seen[h_str] = 1
            if h_str not in ("占比", "环比", ""):
                last_unique = h_str
    return result


def read_sheet(wb, sheet_name):
    """读取指定 Sheet 的所有数据"""
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    data = []
    headers = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        row_data = []
        for cell in row:
            if cell is None:
                row_data.append("")
            elif isinstance(cell, (datetime, date)):
                row_data.append(cell.strftime("%Y-%m-%d"))
            else:
                row_data.append(cell)

        if row_idx == 0:
            raw_headers = [str(h) for h in row_data]
            headers = deduplicate_headers(raw_headers)
        else:
            data.append(dict(zip(headers, row_data)))

    return {"headers": headers, "rows": data, "row_count": len(data)}


def main():
    if len(sys.argv) < 2:
        print("用法: python read_excel.py <工作目录路径> [--sheet <Sheet名>]")
        sys.exit(1)

    work_dir = sys.argv[1]
    target_sheet = None
    output_override = None

    if '--sheet' in sys.argv:
        idx = sys.argv.index('--sheet')
        if idx + 1 < len(sys.argv):
            target_sheet = sys.argv[idx + 1]

    if '--output' in sys.argv:
        idx = sys.argv.index('--output')
        if idx + 1 < len(sys.argv):
            output_override = sys.argv[idx + 1]

    # 查找数据文件
    filepath, cutoff_date, last_week_file = find_latest_gmv_file(work_dir)
    if not filepath:
        print(json.dumps({"error": "未找到 GMV全量数据 文件", "work_dir": work_dir},
                         ensure_ascii=False, indent=2))
        sys.exit(1)

    print(f"📂 数据文件: {os.path.basename(filepath)}", file=sys.stderr)
    print(f"📅 截止日期: {cutoff_date}", file=sys.stderr)
    if last_week_file:
        print(f"📂 上周文件: {os.path.basename(last_week_file)}", file=sys.stderr)

    # 打开 Excel
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"📋 可用 Sheet: {wb.sheetnames}", file=sys.stderr)

    result = {
        "file": os.path.basename(filepath),
        "cutoff_date": cutoff_date,
        "time_progress": calc_time_progress(cutoff_date),
        "available_sheets": wb.sheetnames,
        "data": {}
    }

    # 添加上周文件信息
    if last_week_file:
        result["last_week_file"] = os.path.basename(last_week_file)

    if target_sheet:
        # 只读取指定 Sheet
        sheet_data = read_sheet(wb, target_sheet)
        if sheet_data:
            result["data"][target_sheet] = sheet_data
        else:
            result["error"] = f"Sheet '{target_sheet}' 不存在"
    else:
        # 读取分析所需的核心 Sheet
        # 动态季度 Sheet 名
        tp = calc_time_progress(cutoff_date)
        quarter = tp.get("quarter", 1)

        core_sheets = [
            "周营收-财务口径",
            "周营收-团队口径",
            f"Q{quarter}季度-B标",
            f"Q{quarter}季度-A标",  # 同时读取A标（包含不同预算标准）
            f"Q{quarter}累计同比"  # 动态季度同比 Sheet
        ]

        # 添加当月 Sheet（当月-A标有预算，当月-B标有SMROI，两者都要读）
        month_sheet = get_current_month_sheet_name(cutoff_date)
        if month_sheet:
            core_sheets.append(month_sheet)  # B标（如4月-B标）
            # 同时读取A标（包含预算数据）
            a_sheet = month_sheet.replace("-B标", "-A标")
            if a_sheet in wb.sheetnames:
                core_sheets.append(a_sheet)

        # 添加当月同比 Sheet
        month_yoy_sheet = get_current_month_yoy_sheet_name(cutoff_date)
        if month_yoy_sheet:
            core_sheets.append(month_yoy_sheet)

        for sheet_name in core_sheets:
            sheet_data = read_sheet(wb, sheet_name)
            if sheet_data:
                result["data"][sheet_name] = sheet_data
                print(f"  ✅ {sheet_name}: {sheet_data['row_count']} 行", file=sys.stderr)
            else:
                print(f"  ⚠️ {sheet_name}: 未找到", file=sys.stderr)

        # 读取上周数据（如果存在）
        if last_week_file:
            try:
                wb_last = openpyxl.load_workbook(last_week_file, data_only=True)
                result["last_week_data"] = {}

                # 如果本周文件是脱敏版，需要把上周数据也做同样的名称映射
                # 这样环比计算时名称才能匹配
                is_current_desensitized = os.path.basename(filepath).startswith("desensitized")
                name_mapping = None
                if is_current_desensitized:
                    mapping_path = os.path.join(work_dir, "mapping.json")
                    if os.path.exists(mapping_path):
                        with open(mapping_path, "r", encoding="utf-8") as mf:
                            mapping_data = json.load(mf)
                            name_mapping = mapping_data.get("name_mapping", {})
                        print(f"  🔄 本周为脱敏数据，将对上周数据应用名称映射", file=sys.stderr)

                last_week_sheets = ["周营收-财务口径", "周营收-团队口径"]
                for sheet_name in last_week_sheets:
                    if sheet_name in wb_last.sheetnames:
                        sheet_data = read_sheet(wb_last, sheet_name)
                        if sheet_data and name_mapping:
                            # 对上周数据应用名称映射（原名→脱敏名）
                            for row in sheet_data["rows"]:
                                for key in ("项目组", "季度数据", "月度数据"):
                                    if key in row and row[key]:
                                        orig_name = str(row[key]).strip()
                                        # 尝试直接映射
                                        if orig_name in name_mapping:
                                            row[key] = name_mapping[orig_name]
                                        else:
                                            # 尝试部分匹配（处理"合计"等含有映射名的情况）
                                            for orig, masked in name_mapping.items():
                                                if orig in orig_name:
                                                    row[key] = orig_name.replace(orig, masked)
                                                    break
                        if sheet_data:
                            result["last_week_data"][sheet_name] = sheet_data
                            print(f"  ✅ 上周 {sheet_name}: {sheet_data['row_count']} 行", file=sys.stderr)

                wb_last.close()
            except Exception as e:
                print(f"  ⚠️ 读取上周文件失败: {e}", file=sys.stderr)

    wb.close()

    # 加载 config.yaml 阈值
    config = load_config()
    if config:
        result["config"] = config

    # 输出 JSON
    print(json.dumps(result, ensure_ascii=False, indent=2))

    # === 生成数据快照文件（防编数机制） ===
    snapshot_path = output_override or os.path.join(work_dir, "data_snapshot.md")
    generate_data_snapshot(result, snapshot_path)
    print(f"ℹ️ 当前处理真实数据；如需脱敏请先运行 desensitize.py", file=sys.stderr)


def format_snapshot_value(v, field_name=""):
    """格式化快照值。field_name用于判断是否为SMROI/ROI类指标（不应格式化为百分比）"""
    if isinstance(v, float):
        # SMROI/ROI类指标保持原值格式，不转百分比
        is_roi = any(kw in str(field_name).upper() for kw in ["SMROI", "ROI"])
        if not is_roi and abs(v) < 2 and abs(v) > 0.0001:
            return f"{v:.2%}"
        return f"{v:,.2f}"
    return str(v)


def extract_key_metrics(result):
    """从关键 sheet 中抽取更容易核对的指标索引。"""
    metrics = []
    for sheet_name, sheet_data in result.get("data", {}).items():
        if not sheet_data or not sheet_data.get("rows"):
            continue
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        # 只对常用核心表抽取索引，避免快照过长难查
        if sheet_name not in ["周营收-财务口径", "周营收-团队口径", "Q1季度-B标"]:
            continue

        key_candidates = ["项目组", "项目", "业务线", "名称", "课程", "团队"]
        metric_candidates = ["营收", "GMV", "拉新GMV", "续报GMV", "SMROI", "进度", "达成率"]

        label_key = None
        for k in key_candidates:
            if k in headers:
                label_key = k
                break

        matched_metrics = [h for h in headers if any(m in str(h) for m in metric_candidates)]
        if not label_key or not matched_metrics:
            continue

        for row in rows:
            label = row.get(label_key, "")
            if label in ("", None):
                continue
            for metric in matched_metrics[:6]:
                value = row.get(metric, "")
                if value not in ("", None):
                    metrics.append({
                        "sheet": sheet_name,
                        "label": label,
                        "metric": metric,
                        "value": format_snapshot_value(value, field_name=metric),
                    })
    return metrics


def generate_data_snapshot(result, output_path):
    """
    生成人类可读的数据快照文件
    分析过程中每个数字都必须能在此文件中找到出处
    """
    lines = []
    lines.append(f"# 数据快照 — {result.get('cutoff_date', '未知日期')}")
    lines.append(f"\n> ⚠️ 本文件由 read_excel.py 自动生成，是分析报告的唯一数据来源")
    lines.append(f"> 报告中的每个数字必须能在本文件中找到。找不到的数字 = 编的数字。")
    lines.append(f"> 如需引用数据，优先从下面的【关键指标索引】查找，再回溯到具体 Sheet 表格。")
    lines.append(f"\n**数据文件**: {result.get('file', '未知')}")

    tp = result.get("time_progress", {})
    if tp:
        lines.append(f"\n## 时间进度")
        lines.append(f"- Q1进度: {tp.get('q1_progress_display', 'N/A')}")
        lines.append(f"- 当月进度: {tp.get('month_progress_display', 'N/A')}")

    key_metrics = extract_key_metrics(result)
    if key_metrics:
        lines.append("\n## 关键指标索引（优先查这里）")
        lines.append("| Sheet | 对象 | 指标 | 数值 |")
        lines.append("| --- | --- | --- | --- |")
        for item in key_metrics[:120]:
            lines.append(f"| {item['sheet']} | {item['label']} | {item['metric']} | {item['value']} |")

    for sheet_name, sheet_data in result.get("data", {}).items():
        lines.append(f"\n## {sheet_name}")
        if not sheet_data or not sheet_data.get("rows"):
            lines.append("（无数据）")
            continue

        headers = sheet_data["headers"]
        lines.append(f"\n| {'  |  '.join(str(h) for h in headers)} |")
        lines.append(f"| {'  |  '.join('---' for _ in headers)} |")

        for row in sheet_data["rows"]:
            vals = [format_snapshot_value(row.get(h, ""), field_name=h) for h in headers]
            lines.append(f"| {'  |  '.join(vals)} |")

    # 上周数据快照
    if result.get("last_week_data"):
        lines.append(f"\n---\n# 上周数据（用于计算环比变动）")
        lines.append(f"**上周文件**: {result.get('last_week_file', '未知')}")

        for sheet_name, sheet_data in result["last_week_data"].items():
            lines.append(f"\n## 上周 - {sheet_name}")
            if not sheet_data or not sheet_data.get("rows"):
                continue

            headers = sheet_data["headers"]
            lines.append(f"\n| {'  |  '.join(str(h) for h in headers)} |")
            lines.append(f"| {'  |  '.join('---' for _ in headers)} |")

            for row in sheet_data["rows"]:
                vals = [format_snapshot_value(row.get(h, ""), field_name=h) for h in headers]
                lines.append(f"| {'  |  '.join(vals)} |")

    # ═══ 嵌入JSON数据块（供 generate_report.py 和 validate_report.py 使用） ═══
    json_data = {
        'file': result.get('file', ''),
        'cutoff_date': result.get('cutoff_date', ''),
        'time_progress': result.get('time_progress', {}),
        'data': {},
        'last_week_data': {},
        'key_metrics': key_metrics[:200] if key_metrics else [],
    }
    # Serialize sheet data (compact: only rows, skip raw headers)
    for sheet_name, sheet_data in result.get("data", {}).items():
        if sheet_data and sheet_data.get("rows"):
            json_data['data'][sheet_name] = {
                'headers': sheet_data['headers'],
                'rows': sheet_data['rows'],
            }
    for sheet_name, sheet_data in result.get("last_week_data", {}).items():
        if sheet_data and sheet_data.get("rows"):
            json_data['last_week_data'][sheet_name] = {
                'headers': sheet_data['headers'],
                'rows': sheet_data['rows'],
            }

    lines.append("")
    lines.append("<!-- DATA_JSON_START")
    lines.append(json.dumps(json_data, ensure_ascii=False, default=str))
    lines.append("DATA_JSON_END -->")
    lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"\n📸 数据快照已生成: {os.path.basename(output_path)}", file=sys.stderr)
    print(f"   ⚠️ 分析报告中的每个数字必须来自此文件", file=sys.stderr)
    if key_metrics:
        print(f"   🔎 已生成关键指标索引，便于快速核对", file=sys.stderr)


if __name__ == "__main__":
    main()
