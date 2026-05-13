#!/usr/bin/env python3
"""
经营分析报告生成脚本

从 data_snapshot.md 的 JSON 数据块中读取数据，
自动生成完整的经营分析报告（Markdown格式）。

用法:
    python generate_report.py <data_snapshot.md> [--output <输出文件路径>]
    python generate_report.py <工作目录路径> [--output <输出文件路径>]  # 兼容旧用法

示例:
    python generate_report.py /path/to/data_snapshot.md
    python generate_report.py /path/to/data --output /path/to/report.md
"""

import sys

# Windows编码兼容
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
import io
import os
import json
import yaml
import subprocess
import calendar
from datetime import datetime, timedelta

# Windows 编码处理
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# ============================================================
# 路径工具：获取 Skill 根目录（用于日志等持久化文件）
# ============================================================
_SKILL_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # .../skills/jingying-fenxi

def _log_dir():
    """返回日志目录（skill-root/logs/），不存在则创建"""
    d = os.path.join(_SKILL_ROOT, 'logs')
    os.makedirs(d, exist_ok=True)
    return d

def _usage_path():
    """返回用量追踪文件路径（skill-root/_usage.jsonl）"""
    return os.path.join(_SKILL_ROOT, '_usage.jsonl')


# ============================================================
# 配置加载
# ============================================================

def load_config():
    """加载 config.yaml 阈值参数"""
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "config.yaml")
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)
    return {}

# 全局配置
_CONFIG = load_config()


def get_threshold(key, default):
    """从config获取阈值"""
    thresholds = _CONFIG.get("thresholds", {})
    return thresholds.get(key, default)


# ============================================================
# 辅助函数
# ============================================================

def load_from_snapshot(snapshot_path):
    """从 data_snapshot.md 中提取嵌入的JSON数据块"""
    import re
    with open(snapshot_path, 'r', encoding='utf-8') as f:
        content = f.read()
    match = re.search(r'<!-- DATA_JSON_START\n(.*?)\nDATA_JSON_END -->', content, re.DOTALL)
    if not match:
        return None
    return json.loads(match.group(1))


def load_json_data(work_dir):
    """加载数据。优先从 data_snapshot.md 的JSON块读取，其次从缓存/read_excel.py"""

    # 优先：从 data_snapshot.md 中读取嵌入JSON
    snapshot_path = os.path.join(work_dir, "data_snapshot.md")
    if os.path.exists(snapshot_path):
        data = load_from_snapshot(snapshot_path)
        if data:
            data["_work_dir"] = work_dir
            print(f"📂 从 data_snapshot.md JSON块加载数据", file=sys.stderr)
            return data
        else:
            print(f"⚠️ data_snapshot.md 中无JSON块，尝试其他方式", file=sys.stderr)

    # 其次：从缓存文件
    cache_path = os.path.join(work_dir, "data_full.json")
    if os.path.exists(cache_path):
        with open(cache_path, encoding="utf-8") as f:
            data = json.load(f)
            data["_work_dir"] = work_dir
            return data

    # 最后：调用 read_excel.py
    read_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "read_excel.py")
    if not os.path.exists(read_script):
        read_script = os.path.join(work_dir, "read_excel.py")
    if not os.path.exists(read_script):
        raise FileNotFoundError("read_excel.py not found")

    result = subprocess.run(
        [sys.executable, read_script, work_dir],
        capture_output=True, text=True, encoding="utf-8"
    )
    if result.returncode != 0:
        raise RuntimeError(f"read_excel.py failed: {result.stderr}")

    text = result.stdout
    first_brace = text.find('{')
    if first_brace == -1:
        raise RuntimeError("No JSON found in read_excel.py output")
    text = text[first_brace:]

    with open(cache_path, "w", encoding="utf-8") as f:
        f.write(text)

    data = json.loads(text)
    data["_work_dir"] = work_dir
    return data


def safe_float(v, default=0.0):
    """安全转 float"""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("%", "")
    if s in ("", "-"):
        return default
    try:
        return float(s)
    except:
        return default


def fmt_wan(v, decimals=1):
    """格式化为万元"""
    if v is None:
        return "-"
    return f"{v:,.{decimals}f}"


def fmt_pct(v, decimals=1, signed=True):
    """格式化为百分比"""
    if v is None:
        return "-"
    sgn = "+" if (signed and v >= 0) else ""
    return f"{sgn}{v:.{decimals}f}%"


def fmt_smroi(v, decimals=2):
    """格式化为 SMROI"""
    if v is None:
        return "-"
    return f"{v:.{decimals}f}"


def calc_change_pct(this_week, last_week):
    """计算环比百分比"""
    tw = safe_float(this_week)
    lw = safe_float(last_week)
    if lw == 0:
        return None
    return (tw - lw) / abs(lw) * 100


def safe_div_pct(numerator, denominator, decimals=1):
    """安全除法并格式化为百分比字符串，防ZeroDivisionError"""
    n = safe_float(numerator)
    d = safe_float(denominator)
    if d == 0:
        return "-"
    return f"{n / d * 100:.{decimals}f}%"


def calc_abs_change(this_week, last_week):
    """计算绝对变动"""
    return safe_float(this_week) - safe_float(last_week)


def progress_status(progress_pct, time_progress_pct, threshold=None):
    """判断进度状态：对比Q2/月份标进度 vs 时间进度"""
    if threshold is None:
        # 合理阈值：>5pp为超前，<-5pp为落后，中间为边缘
        threshold = get_threshold("absolute_change", 0.05) * 100  # 默认5pp
        if threshold > 10:  # 如果config里配的值>10pp，改用5pp
            threshold = 5
    diff = safe_float(progress_pct) - safe_float(time_progress_pct)
    if diff > threshold:
        return "✅超前", f"+{diff:.1f}pp"
    elif diff < -threshold:
        return "🔴落后", f"{diff:.1f}pp"
    else:
        return "⚠️边缘", f"{diff:+.1f}pp"


def scan_risks(data, q_progress, month_progress, quarter=1):
    """综合风险扫描：进度落后、SMROI<1、目标达成<70%、GMV负值、周跌幅>50%"""
    risks = []
    fin_rows = get_week_rows(data, "周营收-财务口径")
    last_fin_rows = get_last_week_rows(data, "周营收-财务口径")

    q_sheet = data.get("data", {}).get(f"Q{quarter}季度-B标", {})
    q_rows = q_sheet.get("rows", [])
    q_map = {}
    for row in q_rows:
        label = str(row.get("季度数据", "")).strip().replace("  └ ", "").replace("└ ", "").strip()
        if label:
            q_map[label] = row

    for row in fin_rows:
        name = row.get("项目组", "")
        if not name or "合计" in str(name):
            continue

        # SMROI < 1
        smroi = safe_float(row.get("SMROI", 0))
        if 0 < smroi < 1:
            risks.append(f"🔴 {name}: SMROI={smroi:.2f} < 1，投放亏损")

        # GMV负值
        gmv = safe_float(row.get("总营收", 0))
        if gmv < 0:
            risks.append(f"🔴 {name}: GMV为负值（{gmv:.1f}万）")

        # 周跌幅 > 50%
        lw_row = get_last_by_name(last_fin_rows, name)
        if lw_row:
            lw_gmv = safe_float(lw_row.get("总营收", 0))
            if lw_gmv > 0:
                chg = (gmv - lw_gmv) / lw_gmv * 100
                if chg < -50:
                    risks.append(f"🔴 {name}: 周跌幅{chg:.1f}%，超过50%警戒线")

    # Q{quarter}目标达成 < 70%
    for label, row in q_map.items():
        if "合计" in label or "个项目组" in label:
            continue
        total_gmv = safe_float(row.get("总GMV", 0))
        target = safe_float(row.get("目标", 0))
        if target > 0:
            achieve = total_gmv / target * 100
            if achieve < 70:
                risks.append(f"⚠️ {label}: Q{quarter}达成率仅{achieve:.0f}%，低于70%警戒线")

        # 进度落后
        p_str = str(row.get("进度%", "0%")).replace("%", "")
        progress = safe_float(p_str)
        if q_progress - progress > 10:
            risks.append(f"⚠️ {label}: Q{quarter}进度{progress:.1f}%，落后时间进度{q_progress}%")

    return risks


def get_week_rows(data, sheet_name):
    """获取周数据行"""
    sheet = data.get("data", {}).get(sheet_name, {})
    rows = sheet.get("rows", [])
    result = []
    for row in rows:
        name = row.get("项目组", "")
        if not name or name in ("", None):
            continue
        if "周" in str(name) and ("第" in str(name)):
            continue
        if "合计" in str(name):
            continue
        result.append(row)
    return result


def get_last_week_rows(data, sheet_name):
    """获取上周数据行"""
    lwd = data.get("last_week_data", {})
    sheet = lwd.get(sheet_name, {})
    rows = sheet.get("rows", [])
    result = []
    for row in rows:
        name = row.get("项目组", "")
        if not name:
            continue
        if "合计" in str(name):
            continue
        result.append(row)
    return result


def get_total_row(data, sheet_name):
    """获取合计行"""
    sheet = data.get("data", {}).get(sheet_name, {})
    for row in sheet.get("rows", []):
        if "合计" in str(row.get("项目组", "")):
            return row
    return None


def get_last_week_total(data, sheet_name):
    """获取上周合计行"""
    sheet = data.get("last_week_data", {}).get(sheet_name, {})
    for row in sheet.get("rows", []):
        if "合计" in str(row.get("项目组", "")):
            return row
    return None


def get_last_by_name(last_rows, name):
    """按名称查找上周数据"""
    for r in last_rows:
        if name in str(r.get("项目组", "")):
            return r
    return None


def extract_week_number(cutoff_date):
    """从日期提取周数（使用ISO周历）"""
    dt = datetime.strptime(cutoff_date, "%Y-%m-%d")
    _, week_num, _ = dt.isocalendar()
    return f"W{week_num}", dt


# ============================================================
# 报告生成
# ============================================================

def extract_business_groups(data, quarter=1):
    """从Q{quarter}季度-B标动态提取业务线分组信息。
    
    解析结构如 "升学 (5个项目组)" + "  └ 全国大班" 来确定分组，
    兼容脱敏后的 "业务线A (5个项目组)" + "  └ 项目组A"。
    
    Returns:
        dict with keys:
            shengxue: list of project names in 升学/业务线A
            suyang: list of project names in 素养/业务线B
            shengxue_label: display label for 升学 group (e.g. "升学 (5个项目组)")
            suyang_label: display label for 素养 group
    """
    q1_sheet = data.get("data", {}).get(f"Q{quarter}季度-B标", {})
    q1_rows = q1_sheet.get("rows", [])
    
    # Parse the tree structure from Q1季度-B标
    groups = {}  # label -> [child_names]
    current_group = None
    current_label = None
    
    for row in q1_rows:
        label = str(row.get("季度数据", "")).strip()
        if not label or label == "总计":
            continue
        
        # Group header: contains "个项目组" (e.g. "升学 (5个项目组)" or "业务线A (5个项目组)")
        if "个项目组" in label:
            current_group = []
            current_label = label
            groups[current_label] = current_group
        # Child item: starts with └ or has └
        elif "└" in label and current_group is not None:
            child_name = label.replace("└", "").strip()
            current_group.append(child_name)
    
    # Identify which group is 升学 and which is 素养
    # Strategy: check for known names first, then fall back to order
    shengxue = []
    suyang = []
    shengxue_label = ""
    suyang_label = ""
    
    # Known original names for identification
    known_shengxue = {"全国大班", "云领学", "研习所", "新世界", "强基业务"}
    known_suyang = {"小图灵", "KP", "纵横工作室", "线下店", "Deepthink", "博闻", "思维"}
    
    for label, children in groups.items():
        children_set = set(children)
        if children_set & known_shengxue:
            # Original names - this is 升学
            shengxue = children
            shengxue_label = label
        elif children_set & known_suyang:
            # Original names - this is 素养
            suyang = children
            suyang_label = label
        elif "升学" in label or "业务线A" in label:
            shengxue = children
            shengxue_label = label
        elif "素养" in label or "业务线B" in label:
            suyang = children
            suyang_label = label
    
    # If we still couldn't identify (e.g. fully desensitized with unknown labels),
    # use order: first group = 升学, second = 素养
    if not shengxue and not suyang:
        group_list = list(groups.items())
        if len(group_list) >= 1:
            shengxue_label, shengxue = group_list[0]
        if len(group_list) >= 2:
            suyang_label, suyang = group_list[1]
    elif not shengxue:
        # One matched, assign the other
        for label, children in groups.items():
            if label != suyang_label:
                shengxue = children
                shengxue_label = label
                break
    elif not suyang:
        for label, children in groups.items():
            if label != shengxue_label:
                suyang = children
                suyang_label = label
                break
    
    # Fallback to hardcoded if Q1季度-B标 is empty/missing
    if not shengxue:
        shengxue = ["全国大班", "云领学", "研习所", "新世界", "强基业务"]
        shengxue_label = "升学 (5个项目组)"
    if not suyang:
        suyang = ["小图灵", "KP", "纵横工作室", "线下店", "Deepthink", "博闻", "思维"]
        suyang_label = "素养 (7个项目组)"
    
    return {
        "shengxue": shengxue,
        "suyang": suyang,
        "shengxue_label": shengxue_label,
        "suyang_label": suyang_label,
    }


def generate_report(data):
    """生成完整报告"""
    cutoff_date = data.get("cutoff_date", "")
    week_label, dt = extract_week_number(cutoff_date)

    week_end = dt
    week_start = dt - timedelta(days=6)
    week_start_str = week_start.strftime("%m月%d日").lstrip("0").replace("月0", "月")
    week_end_str = dt.strftime("%m月%d日").lstrip("0").replace("月0", "月")

    tp = data.get("time_progress", {})
    q_progress = tp.get("q_progress", 0)
    quarter = tp.get("quarter", 1)  # 动态季度：1=Q1, 2=Q2, 3=Q3, 4=Q4
    month_progress = tp.get("month_progress", 0)

    today = datetime.now().strftime("%Y-%m-%d")

    # 提取数据
    this_fin = get_week_rows(data, "周营收-财务口径")
    last_fin = get_last_week_rows(data, "周营收-财务口径")
    this_team = get_week_rows(data, "周营收-团队口径")
    last_team = get_last_week_rows(data, "周营收-团队口径")

    this_fin_total = get_total_row(data, "周营收-财务口径")
    last_fin_total = get_last_week_total(data, "周营收-财务口径")
    this_team_total = get_total_row(data, "周营收-团队口径")
    last_team_total = get_last_week_total(data, "周营收-团队口径")

    # 判断是否有上周数据（用于环比和变动额显示）
    has_last_week = bool(last_fin) or bool(last_fin_total)

    # 动态提取业务线分组（兼容脱敏/原始名称）
    biz_groups = extract_business_groups(data, quarter)
    suyang = biz_groups["suyang"]
    shengxue = biz_groups["shengxue"]
    suyang_label = biz_groups["suyang_label"]
    shengxue_label = biz_groups["shengxue_label"]

    # 构建素养名称集合用于快速查找
    suyang_set = set(suyang)

    def is_suyang(name):
        name_str = str(name)
        return any(s in name_str for s in suyang)

    lines = []

    # ---- 标题 ----
    lines.append(f"# {week_label} 经营分析报告（{week_start_str}-{week_end_str}）")
    lines.append("")
    lines.append(f"> 数据截止：{cutoff_date} | 分析日期：{today} | 模式：纯模板生成（模型无关）")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ---- 时间进度基准 ----
    lines.append("## 时间进度基准")
    lines.append("")
    # 动态计算季度天数和当月天数
    q_total_days = tp.get("q_total_days", 90)
    q_elapsed_days = tp.get("q_elapsed_days", 0)
    month_total_days = tp.get("month_total_days", calendar.monthrange(dt.year, dt.month)[1])
    month_elapsed_days = tp.get("month_elapsed_days", dt.day)
    month_names = {1: "1月", 2: "2月", 3: "3月", 4: "4月", 5: "5月", 6: "6月",
                   7: "7月", 8: "8月", 9: "9月", 10: "10月", 11: "11月", 12: "12月"}
    current_month_name = month_names.get(dt.month, f"{dt.month}月")
    lines.append(f"- **Q{quarter}时间进度**：{q_progress}%（{q_elapsed_days}/{q_total_days}天）")
    lines.append(f"- **{current_month_name}时间进度**：{month_progress}%（{month_elapsed_days}/{month_total_days}天）")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ============================================================
    # 模块1：单周营收趋势（财务口径）
    # ============================================================
    lines.append("## 模块1：单周营收趋势（财务口径）")
    lines.append("")
    lines.append(f"**{week_label}（{week_start_str}-{week_end_str}）数据：**")
    lines.append("")

    tw_total = safe_float(this_fin_total.get("总营收") if this_fin_total else 0)
    lw_total = safe_float(last_fin_total.get("总营收") if last_fin_total else 0)
    tw_laixin = safe_float(this_fin_total.get("拉新GMV") if this_fin_total else 0)
    lw_laixin = safe_float(last_fin_total.get("拉新GMV") if last_fin_total else 0)
    tw_xubao = safe_float(this_fin_total.get("续报GMV") if this_fin_total else 0)
    lw_xubao = safe_float(last_fin_total.get("续报GMV") if last_fin_total else 0)
    tw_smroi = safe_float(this_fin_total.get("SMROI") if this_fin_total else 0)
    lw_smroi = safe_float(last_fin_total.get("SMROI") if last_fin_total else 0)

    chg_total = calc_change_pct(tw_total, lw_total)
    chg_laixin = calc_change_pct(tw_laixin, lw_laixin)
    chg_xubao = calc_change_pct(tw_xubao, lw_xubao)
    chg_smroi = calc_change_pct(tw_smroi, lw_smroi)
    abs_total = calc_abs_change(tw_total, lw_total)
    abs_laixin = calc_abs_change(tw_laixin, lw_laixin)
    abs_xubao = calc_abs_change(tw_xubao, lw_xubao)
    abs_smroi = calc_abs_change(tw_smroi, lw_smroi)

    def emoji(chg):
        if chg is None:
            return "-"
        if chg == 0:
            return "➡️"
        return "🟢" if chg > 0 else "🔴"

    def smroi_emoji(this_val, last_val):
        """SMROI趋势颜色判断（v6.6.0）
        规则：负数→🔴 | 从正变负→🔴 | 从负变正→🟢 | 都在正→增加→🟢 | 都在负→绝对值减少→🟢"""
        if this_val is None or last_val is None:
            return "⚪"
        this_f = float(this_val) if not isinstance(this_val, (int, float)) else this_val
        last_f = float(last_val) if not isinstance(last_val, (int, float)) else last_val
        if this_f < 0:
            return "🔴"
        if last_f >= 0 and this_f < 0:
            return "🔴"
        if last_f < 0 and this_f >= 0:
            return "🟢"
        if last_f >= 0 and this_f >= 0:
            return "🟢" if this_f > last_f else "🔴"
        if last_f < 0 and this_f < 0:
            return "🟢" if abs(this_f) < abs(last_f) else "🔴"
        return "⚪"



    def sign(v):
        if v is None:
            return ""
        return "+" if v >= 0 else ""

    lines.append("| 指标      | 本周      | 环比     | 绝对变动 |")
    lines.append("| --------- | --------- | -------- | -------- |")
    if has_last_week:
        lines.append(f"| 总营收    | {fmt_wan(tw_total)}万 | {emoji(chg_total)}{fmt_pct(chg_total)} | {sign(abs_total)}{fmt_wan(abs_total)}万 |")
        lines.append(f"| 拉新GMV   | {fmt_wan(tw_laixin)}万 | {emoji(chg_laixin)}{fmt_pct(chg_laixin)} | {sign(abs_laixin)}{fmt_wan(abs_laixin)}万 |")
        lines.append(f"| 续报GMV   | {fmt_wan(tw_xubao)}万 | {emoji(chg_xubao)}{fmt_pct(chg_xubao)} | {sign(abs_xubao)}{fmt_wan(abs_xubao)}万 |")
        lines.append(f"| 拉新SMROI | {fmt_smroi(tw_smroi)} | {emoji(chg_smroi)}{fmt_pct(chg_smroi)} | {sign(abs_smroi)}{fmt_smroi(abs_smroi)} |")
    else:
        lines.append(f"| 总营收    | {fmt_wan(tw_total)}万 | 无上周数据 | -- |")
        lines.append(f"| 拉新GMV   | {fmt_wan(tw_laixin)}万 | 无上周数据 | -- |")
        lines.append(f"| 续报GMV   | {fmt_wan(tw_xubao)}万 | 无上周数据 | -- |")
        lines.append(f"| 拉新SMROI | {fmt_smroi(tw_smroi)} | 无上周数据 | -- |")
    lines.append("")

    # 趋势描述 — 四象限分析（量=GMV变化，效=SMROI变化）
    if has_last_week:
        gmv_up = chg_total is not None and chg_total > 0
        smroi_up = chg_smroi is not None and chg_smroi > 0
        if gmv_up and smroi_up:
            trend = "量增效增"
        elif gmv_up and not smroi_up:
            trend = "量增效降"
        elif not gmv_up and smroi_up:
            trend = "量跌效升"
        else:
            trend = "量跌效跌"
        lines.append(f"**整体趋势：[待分析]**")
    else:
        lines.append(f"**整体概览** — 总营收{fmt_wan(tw_total)}万，拉新GMV {fmt_wan(tw_laixin)}万，SMROI {fmt_smroi(tw_smroi)}，续报GMV {fmt_wan(tw_xubao)}万。（无上周数据，环比暂不可用）")
    lines.append("")

    # 拉新Top3
    lines.append("### 拉新GMV Top3 及 SMROI 趋势")
    lines.append("")
    lines.append("按拉新GMV绝对变动额排序（本周 - 上周）：")
    lines.append("")

    laixin_top = []
    for row in this_fin:
        name = row.get("项目组", "")
        if not name:
            continue
        tw_val = safe_float(row.get("拉新GMV"))
        lw_row = get_last_by_name(last_fin, name)
        lw_val = safe_float(lw_row.get("拉新GMV") if lw_row else 0)
        smroi_val = row.get("SMROI", "-")
        abs_chg = tw_val - lw_val if has_last_week else None
        laixin_top.append((name, abs_chg, tw_val, smroi_val))

    # 无上周数据时按本周拉新GMV排序；有上周数据时按变动额排序
    if has_last_week:
        laixin_top.sort(key=lambda x: x[1] if x[1] is not None else 0, reverse=True)
    else:
        laixin_top.sort(key=lambda x: x[2], reverse=True)

    lines.append("| 排名 | 项目组 | 拉新GMV变动 | 本周SMROI | 说明 |")
    lines.append("| ---- | -------- | ----------- | --------- | ---- |")
    for i, (name, abs_chg, tw_val, smroi_val) in enumerate(laixin_top[:3], 1):
        smroi_float = safe_float(smroi_val) if str(smroi_val) not in ("", "-", "None") else None
        smroi_str = fmt_smroi(smroi_float) if smroi_float is not None else "-"
        if has_last_week and abs_chg is not None:
            chg_str = f"{emoji(abs_chg)}{sign(abs_chg)}{fmt_wan(abs_chg)}万"
            # 始终尝试查找上周SMROI（先精确匹配项目名）
            lw_row = None
            for r in last_fin:
                if r.get("项目组", "") == name:
                    lw_row = r
                    break
            lw_smroi_val = lw_row.get("SMROI", "-") if lw_row else "-"
            lw_smroi_float = safe_float(lw_smroi_val) if str(lw_smroi_val) not in ("", "-", "None") else None
            # v6.6.0：分析由模型在Step 4填写，此处仅输出占位符
            note = "[待分析]"
        else:
            chg_str = "--"
            note = "[待分析]"
        lines.append(f"| {i} | {name} | {chg_str} | {smroi_str} | {note} |")
    lines.append("")

    # 续报Top3
    lines.append("**续报变动Top3：**")
    lines.append("")
    lines.append("| 排名 | 项目组 | 续报GMV变动 |")
    lines.append("| ---- | -------- | ----------- |")

    xubao_top = []
    for row in this_fin:
        name = row.get("项目组", "")
        if not name:
            continue
        tw_val = safe_float(row.get("续报GMV"))
        lw_row = get_last_by_name(last_fin, name)
        lw_val = safe_float(lw_row.get("续报GMV") if lw_row else 0)
        abs_chg = tw_val - lw_val if has_last_week else None
        xubao_top.append((name, abs_chg, tw_val))

    if has_last_week:
        xubao_top.sort(key=lambda x: x[1] if x[1] is not None else 0, reverse=True)
    else:
        xubao_top.sort(key=lambda x: x[2], reverse=True)
    for i, (name, abs_chg, tw_val) in enumerate(xubao_top[:3], 1):
        if has_last_week and abs_chg is not None:
            chg_str = f"{emoji(abs_chg)}{sign(abs_chg)}{fmt_wan(abs_chg)}万"
        else:
            chg_str = "--"
        lines.append(f"| {i} | {name} | {chg_str} |")

    lines.append("")
    lines.append("> ⚠️ [*说明栏由AI模型补充或手动填写*]")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ============================================================
    # 模块2：素养项目组周营收（团队口径）
    # ============================================================
    lines.append("## 模块2：素养项目组周营收（团队口径）")
    lines.append("")

    # 使用round()防止浮点累加精度问题（如60.400000000000006）
    syang_this = {"总营收": 0.0, "市场GMV": 0.0, "辅导GMV": 0.0}
    for row in this_team:
        if is_suyang(row.get("项目组", "")):
            syang_this["总营收"] = round(syang_this["总营收"] + safe_float(row.get("总营收")), 2)
            syang_this["市场GMV"] = round(syang_this["市场GMV"] + safe_float(row.get("市场GMV")), 2)
            syang_this["辅导GMV"] = round(syang_this["辅导GMV"] + safe_float(row.get("辅导GMV")), 2)

    syang_last = {"总营收": 0.0, "市场GMV": 0.0, "辅导GMV": 0.0}
    for row in last_team:
        if is_suyang(row.get("项目组", "")):
            syang_last["总营收"] = round(syang_last["总营收"] + safe_float(row.get("总营收")), 2)
            syang_last["市场GMV"] = round(syang_last["市场GMV"] + safe_float(row.get("市场GMV")), 2)
            syang_last["辅导GMV"] = round(syang_last["辅导GMV"] + safe_float(row.get("辅导GMV")), 2)

    chg_s_rev = calc_change_pct(syang_this["总营收"], syang_last["总营收"])
    chg_s_mkt = calc_change_pct(syang_this["市场GMV"], syang_last["市场GMV"])
    chg_s_ment = calc_change_pct(syang_this["辅导GMV"], syang_last["辅导GMV"])

    lines.append("**素养整体汇总：**")
    lines.append("")
    lines.append("| 指标 | 本周 | 环比 |")
    lines.append("| ------- | ------- | --------- |")
    lines.append(f"| 总营收 | {fmt_wan(syang_this['总营收'])}万 | {emoji(chg_s_rev)}{fmt_pct(chg_s_rev)} |")
    lines.append(f"| 市场GMV | {fmt_wan(syang_this['市场GMV'])}万 | {emoji(chg_s_mkt)}{fmt_pct(chg_s_mkt)} |")
    lines.append(f"| 辅导GMV | {fmt_wan(syang_this['辅导GMV'])}万 | {emoji(chg_s_ment)}{fmt_pct(chg_s_ment)} |")
    lines.append("")

    lines.append("**各素养项目组详情：**")
    lines.append("")
    lines.append("| 项目组 | 总营收 | 市场GMV | 辅导GMV | SMROI | 环比趋势 |")
    lines.append("| ---------- | ------- | ------- | ------- | ----- | --------- |")

    for sname in suyang:
        for row in this_team:
            if sname in str(row.get("项目组", "")):
                this_rev = safe_float(row.get("总营收"))
                mkt = safe_float(row.get("市场GMV"))
                ment = safe_float(row.get("辅导GMV"))
                smroi_v = row.get("市场SMROI", "-")
                smroi_str = fmt_smroi(safe_float(smroi_v)) if str(smroi_v) not in ("", "-", "None") else "-"
                lw_row = get_last_by_name(last_team, sname)
                lw_rev = safe_float(lw_row.get("总营收")) if lw_row else 0
                lw_smroi_v_raw = lw_row.get("市场SMROI") if lw_row else None
                lw_smroi_v = safe_float(lw_smroi_v_raw) if lw_smroi_v_raw is not None and str(lw_smroi_v_raw) not in ("", "-", "None") else None
                chg = calc_change_pct(this_rev, lw_rev)
                lines.append(f"| {sname} | {fmt_wan(this_rev)}万 | {fmt_wan(mkt)}万 | {fmt_wan(ment)}万 | {smroi_str} | {smroi_emoji(smroi_v, lw_smroi_v)}{fmt_pct(chg)} |")
                break
    lines.append("")
    # Auto-fill素养动因总结
    lines.append("**素养整体动因总结：[待分析]**")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ============================================================
    # 模块3：各项目组GMV进度及SMROI达标
    # 优先级逻辑：优先看B标达成率，B标达成后再对比A标（超额完成部分）
    # ============================================================
    lines.append("## 模块3：各项目组GMV进度及SMROI达标")
    lines.append("")

    # 动态季度Sheet（Q1季度-B标 → Q2季度-B标）
    q_sheet_b = data.get("data", {}).get(f"Q{quarter}季度-B标", {})
    q_sheet_a = data.get("data", {}).get(f"Q{quarter}季度-A标", {})
    q_rows_b = q_sheet_b.get("rows", [])
    q_rows_a = q_sheet_a.get("rows", [])
    q_map_b = {}
    q_map_a = {}
    for row in q_rows_b:
        label = str(row.get("季度数据", "")).strip().replace("  └ ", "").replace("└ ", "").strip()
        if label:
            q_map_b[label] = row
    for row in q_rows_a:
        label = str(row.get("季度数据", "")).strip().replace("  └ ", "").replace("└ ", "").strip()
        if label:
            q_map_a[label] = row

    lines.append(f"### Q{quarter}季度进度分析（vs {q_progress}%时间进度）")
    lines.append("")
    # 三层进度（总进度 + 拉新进度 + 续报进度）+ A标/B标双列对比
    lines.append(f"| 项目组 | Q{quarter}进度(B标) | 状态 | B标总达成率 | A标总达成率 | B标拉新进度 | B标续报进度 |")
    lines.append("| ---------- | ------- | -------- | ------ | ------ | ------ | ------ |")

    all_projects = shengxue + suyang
    q_items = []
    for proj in all_projects:
        row_b = None
        row_a = None
        for label, row in q_map_b.items():
            label_clean = label.replace("└", "").strip()
            if proj == label_clean or proj in label:
                row_b = row
                break
        for label, row in q_map_a.items():
            label_clean = label.replace("└", "").strip()
            if proj == label_clean or proj in label:
                row_a = row
                break
        if row_b is None and row_a is None:
            continue
        # 进度/SMROI从B标取，预算优先B标（B标为空则用A标）
        row_use = row_b if row_b else row_a
        p_str = str(row_use.get("进度%", "0%")).replace("%", "")
        progress = safe_float(p_str)
        status, gap = progress_status(progress, q_progress)
        # SMROI
        smroi_v = row_use.get("拉新SMROI", "-")
        smroi_str = fmt_smroi(safe_float(smroi_v)) if str(smroi_v) not in ("", "None") else "-"
        # B标达成率（优先用B标预算）
        total_b = safe_float(row_b.get("总GMV")) if row_b else 0
        target_b = safe_float(row_b.get("目标")) if row_b else 0
        dacheng_b = (total_b / target_b * 100) if target_b else 0
        dacheng_b_str = f"✅{dacheng_b:.0f}%" if dacheng_b >= 100 else f"🔴{dacheng_b:.0f}%"
        # A标达成率（用于超额对比）
        total_a = safe_float(row_a.get("总GMV")) if row_a else 0
        target_a = safe_float(row_a.get("目标")) if row_a else 0
        dacheng_a = (total_a / target_a * 100) if target_a else 0
        dacheng_a_str = f"✅{dacheng_a:.0f}%" if dacheng_a >= 100 else f"🔴{dacheng_a:.0f}%"
        # 拉新进度（B标）
        laixin_pct = row_b.get("拉新进度", "-") if row_b else "-"
        laixin_pct_str = str(laixin_pct).replace("%", "") if laixin_pct != "-" else "-"
        # 续报进度（B标）
        xubao_pct = row_b.get("续报进度", "-") if row_b else "-"
        xubao_pct_str = str(xubao_pct).replace("%", "") if xubao_pct != "-" else "-"
        q_items.append((proj, progress, status, gap, smroi_str, dacheng_b_str, dacheng_a_str, laixin_pct_str, xubao_pct_str))

    q_items.sort(key=lambda x: x[1], reverse=True)
    for proj, progress, status, gap, smroi_str, dacheng_b_str, dacheng_a_str, laixin_pct_str, xubao_pct_str in q_items:
        lines.append(f"| {proj} | {progress:.1f}% | {status} | {dacheng_b_str} | {dacheng_a_str} | {laixin_pct_str}% | {xubao_pct_str}% |")

    lines.append("")

    # 当月进度（动态确定sheet名）
    # 注意：预算数据在"4月-A标"，"4月-B标"无预算故用于显示进度
    month_sheet_name_a = f"{current_month_name}-A标"
    month_sheet_name_b = f"{current_month_name}-B标"
    month_sheet_a = data.get("data", {}).get(month_sheet_name_a, {})
    month_sheet_b = data.get("data", {}).get(month_sheet_name_b, {})
    month_rows_a = month_sheet_a.get("rows", [])
    month_rows_b = month_sheet_b.get("rows", [])
    # 预算数据从A标取，SMROI/进度从B标取
    month_map_a = {}
    month_map_b = {}
    for row in month_rows_a:
        label = str(row.get("月度数据", "")).strip().replace("  └ ", "").replace("└ ", "").strip()
        if label:
            month_map_a[label] = row
    for row in month_rows_b:
        label = str(row.get("月度数据", "")).strip().replace("  └ ", "").replace("└ ", "").strip()
        if label:
            month_map_b[label] = row

    lines.append(f"### {current_month_name}进度分析（vs {month_progress}%时间进度）")
    lines.append("")
    lines.append(f"| 项目组 | {current_month_name}进度(B标) | 状态 | B标总达成率 | A标总达成率 | B标拉新进度 | B标续报进度 |")
    lines.append("| ---------- | ------- | -------- | ------ | ------ | ------ | ------ |")

    month_items = []
    for proj in all_projects:
        row_a = None
        row_b = None
        for label, row in month_map_a.items():
            label_clean = label.replace("└", "").strip()
            if proj == label_clean or proj in label:
                row_a = row
                break
        for label, row in month_map_b.items():
            label_clean = label.replace("└", "").strip()
            if proj == label_clean or proj in label:
                row_b = row
                break
        if row_a is None and row_b is None:
            continue
        # 进度/SMROI从B标取，预算优先B标（B标为空则用A标）
        row_use = row_b if row_b else row_a
        p_str = str(row_use.get("进度%", "0%")).replace("%", "")
        progress = safe_float(p_str)
        status, gap = progress_status(progress, month_progress)
        smroi_v = row_use.get("拉新SMROI", "-")
        smroi_str = fmt_smroi(safe_float(smroi_v)) if str(smroi_v) not in ("", "None") else "-"
        # B标达成率
        total_b = safe_float(row_b.get("总GMV")) if row_b else 0
        target_b = safe_float(row_b.get("目标")) if row_b else 0
        dacheng_b = (total_b / target_b * 100) if target_b else 0
        dacheng_b_str = f"✅{dacheng_b:.0f}%" if dacheng_b >= 100 else f"🔴{dacheng_b:.0f}%"
        # A标达成率
        total_a = safe_float(row_a.get("总GMV")) if row_a else 0
        target_a = safe_float(row_a.get("目标")) if row_a else 0
        dacheng_a = (total_a / target_a * 100) if target_a else 0
        dacheng_a_str = f"✅{dacheng_a:.0f}%" if dacheng_a >= 100 else f"🔴{dacheng_a:.0f}%"
        # 拉新进度/续报进度
        laixin_pct = row_b.get("拉新进度", "-") if row_b else "-"
        laixin_pct_str = str(laixin_pct).replace("%", "") if laixin_pct != "-" else "-"
        xubao_pct = row_b.get("续报进度", "-") if row_b else "-"
        xubao_pct_str = str(xubao_pct).replace("%", "") if xubao_pct != "-" else "-"
        month_items.append((proj, progress, status, gap, smroi_str, dacheng_b_str, dacheng_a_str, laixin_pct_str, xubao_pct_str))

    month_items.sort(key=lambda x: x[1], reverse=True)
    for proj, progress, status, gap, smroi_str, dacheng_b_str, dacheng_a_str, laixin_pct_str, xubao_pct_str in month_items:
        lines.append(f"| {proj} | {progress:.1f}% | {status} | {dacheng_b_str} | {dacheng_a_str} | {laixin_pct_str}% | {xubao_pct_str}% |")

    lines.append("")
    lines.append("---")
    lines.append("")

    # ============================================================
    # 模块4：与上周对比（复利跟踪）
    # ============================================================
    lines.append("## 模块4：与上周对比（复利跟踪）")
    lines.append("")
    lines.append("### 风险等级变化追踪")
    lines.append("")
    lines.append("> ⚠️ **本模块需要画像卡数据支持**，当前版本暂无上周画像卡数据，风险等级变化表需手动填写。")
    lines.append("")
    lines.append("[待补充-画像卡数据]")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ============================================================
    # 模块5：Q{quarter}累计同比分析
    # ============================================================
    lines.append(f"## 模块5：Q{quarter}累计同比分析")
    lines.append("")
    lines.append("### 整体同比表现")
    lines.append("")

    yoy_sheet = data.get("data", {}).get(f"Q{quarter}累计同比", {})
    yoy_rows = yoy_sheet.get("rows", [])

    def get_yoy(rows, metric):
        for row in rows:
            if str(row.get("季度同比数据", "")).strip() == metric:
                return row.get("_col_1"), row.get("_col_2"), row.get("_col_3")
        return None, None, None

    yoy_total_26, yoy_total_25, yoy_total_chg = get_yoy(yoy_rows, "总GMV")
    yoy_laixin_26, yoy_laixin_25, yoy_laixin_chg = get_yoy(yoy_rows, "拉新GMV")
    yoy_xubao_26, yoy_xubao_25, yoy_xubao_chg = get_yoy(yoy_rows, "续报GMV")
    yoy_smroi_26, yoy_smroi_25, yoy_smroi_chg = get_yoy(yoy_rows, "平均SMROI")

    lines.append(f"| 指标 | 2026年Q{quarter} | 2025年Q{quarter} | 同比变化 |")
    lines.append("| --------- | -------- | -------- | -------- |")
    lines.append(f"| 总GMV | {fmt_wan(safe_float(yoy_total_26))}万 | {fmt_wan(safe_float(yoy_total_25))}万 | {yoy_total_chg} |")
    lines.append(f"| 拉新GMV | {fmt_wan(safe_float(yoy_laixin_26))}万 | {fmt_wan(safe_float(yoy_laixin_25))}万 | {yoy_laixin_chg} |")
    lines.append(f"| 续报GMV | {fmt_wan(safe_float(yoy_xubao_26))}万 | {fmt_wan(safe_float(yoy_xubao_25))}万 | {yoy_xubao_chg} |")
    lines.append(f"| 平均SMROI | {yoy_smroi_26} | {yoy_smroi_25} | {yoy_smroi_chg} |")
    lines.append("")

    # 业务线同比
    lines.append("### 业务线同比分析")
    lines.append("")

    def get_yoy_biz(rows, biz_name):
        for row in rows:
            label = str(row.get("季度同比数据", "")).strip()
            if biz_name in label and "项目组" not in str(row.get("_col_1", "")):
                return row
        return None

    # Extract short group names for YoY lookup (e.g. "升学" from "升学 (5个项目组)" or "业务线A" from "业务线A (5个项目组)")
    shengxue_short = shengxue_label.split("(")[0].split("（")[0].strip() if shengxue_label else "升学"
    suyang_short = suyang_label.split("(")[0].split("（")[0].strip() if suyang_label else "素养"

    sx_row = get_yoy_biz(yoy_rows, shengxue_short)
    sy_row = get_yoy_biz(yoy_rows, suyang_short)

    lines.append(f"**{shengxue_short}业务线（{len(shengxue)}个项目组）：**")
    lines.append("")
    if sx_row:
        lines.append(f"- 总GMV：{fmt_wan(safe_float(sx_row.get('_col_3')))}万 vs {fmt_wan(safe_float(sx_row.get('_col_4')))}万（{sx_row.get('_col_5')}）")
        lines.append(f"- 拉新GMV：{fmt_wan(safe_float(sx_row.get('_col_6')))}万 vs {fmt_wan(safe_float(sx_row.get('_col_7')))}万（{sx_row.get('_col_8')}）")
        lines.append(f"- SMROI：{sx_row.get('_col_9')} vs {sx_row.get('_col_10')}（{sx_row.get('_col_11')}）")
    lines.append("")

    lines.append(f"**{suyang_short}业务线（{len(suyang)}个项目组）：**")
    lines.append("")
    if sy_row:
        lines.append(f"- 总GMV：{fmt_wan(safe_float(sy_row.get('_col_3')))}万 vs {fmt_wan(safe_float(sy_row.get('_col_4')))}万（{sy_row.get('_col_5')}）")
        lines.append(f"- 拉新GMV：{fmt_wan(safe_float(sy_row.get('_col_6')))}万 vs {fmt_wan(safe_float(sy_row.get('_col_7')))}万（{sy_row.get('_col_8')}）")
        lines.append(f"- SMROI：{sy_row.get('_col_9')} vs {sy_row.get('_col_10')}（{sy_row.get('_col_11')}）")
    lines.append("")

    # 亮点和挑战项目组
    lines.append("### 重点项目组同比亮点")
    lines.append("")
    lines.append("| 项目组 | 2026 GMV | 2025 GMV | 同比增长 | 关键表现 |")
    lines.append("| -------- | -------- | -------- | --------- | ---------------------- |")

    proj_yoy = []
    for row in yoy_rows:
        label = str(row.get("季度同比数据", "")).strip()
        if "└" in label and "项目组" in str(row.get("_col_1", "")):
            name = label.replace("└", "").strip()
            gmv_26 = safe_float(row.get("_col_3"))
            gmv_25 = safe_float(row.get("_col_4"))
            chg_str = str(row.get("_col_5", ""))
            chg_val = safe_float(chg_str.replace("%", "").replace("+", ""))
            if "-" in chg_str:
                chg_val = -chg_val
            proj_yoy.append((name, gmv_26, gmv_25, chg_str, chg_val))

    # 亮点（正增长）
    highlights = sorted([p for p in proj_yoy if p[4] > 0], key=lambda x: x[4], reverse=True)[:5]
    for name, gmv_26, gmv_25, chg_str, chg_val in highlights:
        pct = abs(chg_val) / abs(gmv_25) * 100 if gmv_25 else 0
        if pct > 50:
            note = "大幅增长"
        elif pct > 20:
            note = "显著增长"
        else:
            note = "稳健增长"
        lines.append(f"| {name} | {fmt_wan(gmv_26)}万 | {fmt_wan(gmv_25)}万 | 🟢{chg_str} | {note} |")

    lines.append("")
    lines.append("### 同比挑战项目组")
    lines.append("")
    lines.append("| 项目组 | 2026 GMV | 2025 GMV | 同比下降 | 主要挑战 |")
    lines.append("| ---------- | -------- | -------- | -------- | -------------------------- |")

    # 挑战（负增长）
    challenges = sorted([p for p in proj_yoy if p[4] < 0], key=lambda x: x[4])[:5]
    for name, gmv_26, gmv_25, chg_str, chg_val in challenges:
        # v6.6.0：分析由模型在Step 4填写，此处仅输出占位符
        note = "[待分析]"
        lines.append(f"| {name} | {fmt_wan(gmv_26)}万 | {fmt_wan(gmv_25)}万 | 🔴{chg_str} | {note} |")

    lines.append("")
    lines.append("---")
    lines.append("")

    # ============================================================
    # 模块6：课程业务（LCT）分析
    # ============================================================
    lines.append("## 模块6：课程业务（LCT）分析")
    lines.append("")
    lines.append(f"**更新人：___ 更新日期：{today}**")
    lines.append("")
    lines.append("### 1. 本周速览")
    lines.append("")
    lines.append(f"| **指标** | **本周实际** | **上周实际** | **环比** | **Q{quarter}累计** | **Q{quarter}预算** | **达成率** |")
    lines.append("| :------------- | :----------------- | :----------------- | :------------- | :--------------- | :--------------- | :--------------- |")

    # 从Q{quarter}季度-B标获取累计数据
    total_qb = q_map_b.get("总计", {})
    q_gmv = safe_float(total_qb.get("总GMV"))
    q_target = safe_float(total_qb.get("目标"))
    q_laixin = safe_float(total_qb.get("拉新QTD"))
    q_laixin_target = safe_float(total_qb.get("预算"))
    q_xubao = safe_float(total_qb.get("续费QTD"))
    q_xubao_target = safe_float(total_qb.get("续费QTD_预算"))

    lines.append(f"| GMV（万） | {fmt_wan(tw_total)} | {fmt_wan(lw_total)} | {fmt_pct(chg_total)} | {fmt_wan(q_gmv)} | {fmt_wan(q_target)} | {safe_div_pct(q_gmv, q_target)} |")
    lines.append(f"| 其中：拉新 | {fmt_wan(tw_laixin)} | {fmt_wan(lw_laixin)} | {fmt_pct(chg_laixin)} | {fmt_wan(q_laixin)} | {fmt_wan(q_laixin_target)} | {safe_div_pct(q_laixin, q_laixin_target)} |")
    lines.append(f"| 其中：续报 | {fmt_wan(tw_xubao)} | {fmt_wan(lw_xubao)} | {fmt_pct(chg_xubao)} | {fmt_wan(q_xubao)} | {fmt_wan(q_xubao_target)} | {safe_div_pct(q_xubao, q_xubao_target)} |")

    # 市场SMROI
    tw_mkt_smroi = safe_float(this_team_total.get("市场SMROI") if this_team_total else 0)
    lw_mkt_smroi = safe_float(last_team_total.get("市场SMROI") if last_team_total else 0)
    chg_mkt_smroi = calc_change_pct(tw_mkt_smroi, lw_mkt_smroi)
    # 市场SMROI Q{quarter}累计值应来自Q{quarter}累计同比（而非Q{quarter}季度-B标）
    # 修复v6.4.1 bug: _col_1是指标名("平均SMROI")，_col_2才是2026年Q{quarter}的SMROI值
    yoy_smroi_q1 = "-"
    yoy_sheet_data = data.get("data", {}).get(f"Q{quarter}累计同比", {})
    for yoy_r in yoy_sheet_data.get("rows", []):
        if "SMROI" in str(yoy_r.get("季度同比数据", "")):
            yoy_smroi_q1 = yoy_r.get("_col_2", "-")
            break
    sf = safe_float
    smroi_target = sf(total_qb.get('SMROI目标', 0))
    smroi_q1_val = sf(yoy_smroi_q1) if yoy_smroi_q1 not in ('-', '', None) else 0
    if smroi_target and smroi_q1_val:
        smroi_diff = smroi_q1_val - smroi_target
        smroi_achieve = f"{smroi_diff:+.2f}"
    else:
        smroi_achieve = "-"
    lines.append(f"| 市场SMROI | {fmt_smroi(tw_mkt_smroi)} | {fmt_smroi(lw_mkt_smroi)} | {fmt_pct(chg_mkt_smroi)} | {yoy_smroi_q1} | {total_qb.get('SMROI目标', '-')} | {smroi_achieve} |")
    lines.append(f"| 项目SMROI | {fmt_smroi(tw_smroi)} | {fmt_smroi(lw_smroi)} | {fmt_pct(chg_smroi)} | {yoy_smroi_q1} | {total_qb.get('SMROI目标', '-')} | {smroi_achieve} |")
    lines.append("")
    # 动态获取全国大班本周总营收（用于计算集中度）和升学SMROI同比
    dqb_this_rev = 0.0
    dqb_smroi_yoy_str = "数据暂不可比"
    if yoy_rows:
        sx_yoy_r = get_yoy_biz(yoy_rows, shengxue_short)
        if sx_yoy_r:
            sm_26 = safe_float(sx_yoy_r.get('_col_9', 0))
            sm_25 = safe_float(sx_yoy_r.get('_col_10', 0))
            if sm_25 != 0:
                dqb_smroi_yoy = sm_26 - sm_25
                dqb_smroi_yoy_str = f"{dqb_smroi_yoy:+.2f}"
    for r in this_fin:
        if '全国大班' in str(r.get('项目组', '')):
            dqb_this_rev = safe_float(r.get('总营收', 0))
            break
    dqb_share = dqb_this_rev / tw_total * 100 if tw_total else 0
    lines.append(f"**一句话总结：[待分析]**")
    lines.append("")

    lines.append("### 2. 分产品线")
    lines.append("")
    lines.append(f"| **产品线** | **本周GMV（万）** | **周环比** | **Q{quarter}累计（万）** | **Q{quarter}预算达成率** | **上年同期（万）** | **季度同比** | **备注** |")
    lines.append("| :--------------- | :---------------------- | :---------------- | :--------------------- | :--------------------- | :------------------ | :------------- | :------------------------------- |")

    # 获取同比数据
    yoy_sheet = data.get("data", {}).get(f"Q{quarter}累计同比", {})
    yoy_rows = yoy_sheet.get("rows", [])
    yoy_map = {}
    for yoy_r in yoy_rows:
        label = str(yoy_r.get("季度同比数据", "")).strip().replace("  └ ", "").replace("└ ", "").strip()
        if label:
            yoy_map[label] = yoy_r

    def get_yoy_info(name):
        """获取同比数据：上年同期GMV和同比变化"""
        yoy_r = yoy_map.get(name, {})
        last_gmv = safe_float(yoy_r.get("_col_4", 0))  # 2025年Q1
        chg = yoy_r.get("_col_5", "-")  # 同比变化
        return last_gmv, chg

    # 升学汇总 - use exact name match from dynamic group
    shengxue_set = set(shengxue)
    sx_this_total = 0
    sx_last_total = 0
    for row in this_team:  # 使用团队口径
        pname = str(row.get("项目组", "")).strip()
        if pname in shengxue_set or any(s in pname for s in shengxue):
            sx_this_total += safe_float(row.get("总营收"))
    for row in last_team:
        pname = str(row.get("项目组", "")).strip()
        if pname in shengxue_set or any(s in pname for s in shengxue):
            sx_last_total += safe_float(row.get("总营收"))

    sx_chg = calc_change_pct(sx_this_total, sx_last_total)
    sx_q1_row = q_map_b.get(shengxue_label, {})
    sx_q1_gmv = safe_float(sx_q1_row.get("总GMV"))
    sx_q1_target = safe_float(sx_q1_row.get("目标"))
    sx_q1_rate = safe_div_pct(sx_q1_gmv, sx_q1_target)
    sx_yoy_last, sx_yoy_chg = get_yoy_info(shengxue_short)

    lines.append(f"| **{shengxue_short}** | **{fmt_wan(sx_this_total)}** | **{fmt_pct(sx_chg)}** | **{fmt_wan(sx_q1_gmv)}** | **{sx_q1_rate}** | **{fmt_wan(sx_yoy_last)}** | **{sx_yoy_chg}** | - |")

    # 各升学项目组（使用团队口径）
    for proj in shengxue:
        for row in this_team:
            if proj in str(row.get("项目组", "")):
                this_rev = safe_float(row.get("总营收"))
                lw_row = get_last_by_name(last_team, proj)
                lw_rev = safe_float(lw_row.get("总营收") if lw_row else 0)
                lw_smroi_v = safe_float((lw_row.get("市场SMROI") if lw_row else None))
                chg = calc_change_pct(this_rev, lw_rev)
                qb_row = q_map_b.get(proj, {})
                if not qb_row:
                    for k, v in q_map_b.items():
                        if proj in k:
                            qb_row = v
                            break
                qb_gmv_p = safe_float(qb_row.get("总GMV"))
                qb_tgt_p = safe_float(qb_row.get("目标"))
                qb_rate_p = safe_div_pct(qb_gmv_p, qb_tgt_p)
                yoy_last, yoy_chg = get_yoy_info(proj)
                lines.append(f"| └ {proj} | {fmt_wan(this_rev)} | {fmt_pct(chg)} | {fmt_wan(qb_gmv_p)} | {qb_rate_p} | {fmt_wan(yoy_last)} | {yoy_chg} | - |")
                break

    # 素养汇总
    sy_q1_row = q_map_b.get(suyang_label, {})
    sy_q1_gmv = safe_float(sy_q1_row.get("总GMV"))
    sy_q1_target = safe_float(sy_q1_row.get("目标"))
    sy_q1_rate = safe_div_pct(sy_q1_gmv, sy_q1_target)
    sy_yoy_last, sy_yoy_chg = get_yoy_info(suyang_short)

    sy_q1_rate_val = safe_float(str(sy_q1_rate).replace('%', '').replace('🟢', '').replace('🔴', '')) if sy_q1_rate not in ('-', '', None) else 0
    if sy_q1_rate_val >= 100:
        sy_note = "Q2达标"
    elif sy_q1_rate_val >= 80:
        sy_note = "接近达标"
    else:
        sy_note = "Q2落后"
    lines.append(f"| **{suyang_short}** | **{fmt_wan(syang_this['总营收'])}** | **{fmt_pct(chg_s_rev)}** | **{fmt_wan(sy_q1_gmv)}** | **{sy_q1_rate}** | **{fmt_wan(sy_yoy_last)}** | **{sy_yoy_chg}** | **{sy_note}** |")

    # 各素养项目组（使用团队口径）
    for proj in suyang:
        for row in this_team:
            if proj in str(row.get("项目组", "")):
                this_rev = safe_float(row.get("总营收"))
                lw_row = get_last_by_name(last_team, proj)
                lw_rev = safe_float(lw_row.get("总营收") if lw_row else 0)
                lw_smroi_v = safe_float((lw_row.get("市场SMROI") if lw_row else None))
                chg = calc_change_pct(this_rev, lw_rev)
                qb_row = None
                for k, v in q_map_b.items():
                    if proj in k:
                        qb_row = v
                        break
                if qb_row:
                    qb_gmv_p = safe_float(qb_row.get("总GMV"))
                    qb_tgt_p = safe_float(qb_row.get("目标"))
                    qb_rate_p = safe_div_pct(qb_gmv_p, qb_tgt_p)
                else:
                    qb_gmv_p = 0
                    qb_rate_p = "-"
                yoy_last, yoy_chg = get_yoy_info(proj)
                lines.append(f"| └ {proj} | {fmt_wan(this_rev)} | {fmt_pct(chg)} | {fmt_wan(qb_gmv_p)} | {qb_rate_p} | {fmt_wan(yoy_last)} | {yoy_chg} | - |")
                break

    lines.append("")
    lines.append("### 3. 关键洞察")
    lines.append("")
    # Auto-fill LCT关键洞察
    # Find升学and素养shares
    shengxuan_rev = sum(safe_float(r.get('总营收', 0)) for r in this_fin if r.get('项目组') in ['全国大班','云领学','强基业务','研习所','新世界'])
    suyang_rev = safe_float(syang_this['总营收'])
    total_rev = tw_total
    shengxuan_share = shengxuan_rev / total_rev * 100 if total_rev else 0
    suyang_share = suyang_rev / total_rev * 100 if total_rev else 0
    top3_yoy = sorted([p for p in proj_yoy if p[4] > 0], key=lambda x: x[4], reverse=True)[:3]
    top3_names = [p[0] for p in top3_yoy]
    # 动态获取升学SMROI同比变化（从Q{quarter}累计同比，yoy_rows已定义）
    sx_smroi_yoy_str = "数据暂不可比"
    sy_smroi_yoy_str = "数据暂不可比"
    if yoy_rows:
        sx_yoy_r = get_yoy_biz(yoy_rows, shengxue_short)
        if sx_yoy_r:
            sm_26 = safe_float(sx_yoy_r.get('_col_9', 0))
            sm_25 = safe_float(sx_yoy_r.get('_col_10', 0))
            if sm_25 != 0:
                sx_smroi_yoy = sm_26 - sm_25
                sx_smroi_yoy_str = f"{sx_smroi_yoy:+.2f}"
        sy_yoy_r = get_yoy_biz(yoy_rows, suyang_short)
        if sy_yoy_r:
            sy_smroi_yoy_str = sy_yoy_r.get('_col_5', '数据暂不可比')

    # ===== 动态找各维度最优/最差项目 =====
    suyang_names2 = ['Deepthink', 'KP', '博闻', '小图灵', '思维', '纵横工作室', '线下店']
    suyang_smroi_list = []
    for r in this_fin:
        name = str(r.get('项目组', ''))
        if name in suyang_names2:
            sv = r.get('SMROI', '-')
            smroi_num = safe_float(sv) if str(sv) not in ('', '-', 'None') else None
            if smroi_num is not None:
                suyang_smroi_list.append((name, smroi_num))
    best_s_name, best_s_smroi = ('暂无', '-')
    worst_s_name, worst_s_smroi = ('暂无', '-')
    if suyang_smroi_list:
        best_s_name, best_s_smroi = max(suyang_smroi_list, key=lambda x: x[1])
        valid = [(n, v) for n, v in suyang_smroi_list if v >= 0]
        if valid:
            worst_s_name, worst_s_smroi = min(valid, key=lambda x: x[1])

    # 动态找续报达成率最高的项目
    best_xb_name, best_xb_pct = '暂无', '-'
    qb_sheet_data = data.get('data', {}).get(f'Q{quarter}季度-B标', {})
    for row in qb_sheet_data.get('rows', []):
        pname = str(row.get('季度数据', '')).strip()
        if not pname or pname in ('总计', f'升学 (5个项目组)', f'素养 (7个项目组)'):
            continue
        xpct = str(row.get('续报进度', '-')).replace('%', '')
        if xpct not in ('-', '', 'None'):
            pct_val = safe_float(xpct)
            if best_xb_pct == '-' or pct_val > best_xb_pct:
                # Strip tree chars: ├ (251c) └ (2514) ─ (2500) ┬ (252c) ┴ (2534) │ (2502)
                TREE_CHARS = '├└─│┬┴┼┤┌┐└┘\u2500-\u257F\s'
                best_xb_name = pname
                for c in list('├└─│┬┴┼┤ ┌┐'):
                    best_xb_name = best_xb_name.replace(c, '')
                best_xb_name = best_xb_name.strip()
                best_xb_pct = pct_val

    pos_signal = '效率提升' if safe_float(sx_smroi_yoy_str) > 0 else '需关注'
    lines.append("**🟢 积极信号：[待分析]**")
    lines.append("**🔴 风险关注：[待分析]**")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ============================================================
    # 模块7：周报模板（Wording自动填充）
    # ============================================================
    lines.append("## 模块7：周报模板（自动填充）")
    lines.append("")
    lines.append("### 整体 Wording 模板")
    lines.append("")
    lines.append("```")

    # 拉新Top3
    top3_laixin = laixin_top[:3]
    laixin_wording = "；".join([
        f"{i}-{name}{emoji(abs_chg)}{sign(abs_chg)}{fmt_wan(abs_chg)}万（SMROI {fmt_smroi(safe_float(smroi))}）"
        for i, (name, abs_chg, tw_val, smroi) in enumerate(top3_laixin, 1)
    ])

    top3_xubao = xubao_top[:3]
    xubao_wording = "；".join([
        f"{i}-{name}{emoji(abs_chg)}{sign(abs_chg)}{fmt_wan(abs_chg)}万"
        for i, (name, abs_chg, tw_val) in enumerate(top3_xubao, 1)
    ])

    lines.append(f"最近一周（{week_label}: {week_start_str}-{week_end_str}）总营收 {fmt_wan(tw_total)}万（环比{emoji(chg_total)}{fmt_pct(chg_total)}），拉新GMV {fmt_wan(tw_laixin)}万（环比{emoji(chg_laixin)}{fmt_pct(chg_laixin)}），续报GMV {fmt_wan(tw_xubao)}万（环比{emoji(chg_xubao)}{fmt_pct(chg_xubao)}），拉新SMROI {fmt_smroi(tw_smroi)}（环比{emoji(chg_smroi)}{fmt_pct(chg_smroi)}）。本周增长主要由以下项目组带动：")
    lines.append("")
    lines.append(f"其中拉新：{laixin_wording}。")
    lines.append("")
    lines.append(f"续报：{xubao_wording}。")
    lines.append("```")
    lines.append("")

    lines.append(f"### {suyang_short} Wording 模板")
    lines.append("")
    lines.append("```")
    # ===== 定性描述 sentence（增长结构，阈值0.5%排除微小波动）=====
    THRESHOLD = 0.5
    chg_s_mkt_val = safe_float(str(chg_s_mkt).replace('%', '').replace('+', '')) if str(chg_s_mkt) not in ('-', '', None, 'nan') else 0
    chg_s_ment_val = safe_float(str(chg_s_ment).replace('%', '').replace('+', '')) if str(chg_s_ment) not in ('-', '', None, 'nan') else 0
    mkt_up = chg_s_mkt_val > THRESHOLD
    mkt_down = chg_s_mkt_val < -THRESHOLD
    ment_up = chg_s_ment_val > THRESHOLD
    ment_down = chg_s_ment_val < -THRESHOLD

    定性 = "[待分析]"

    # ===== 找最大市场GMV项目组作为核心驱动 =====
    # 注意：suyang_this已在模块2计算，此处复用（使用round防浮点累加问题）
    suyang_this_for_wording = {'总营收': 0.0, '市场GMV': 0.0, '辅导GMV': 0.0}
    for row in this_team:
        name = str(row.get('项目组', ''))
        if any(s in name for s in suyang):
            suyang_this_for_wording['总营收'] = round(suyang_this_for_wording['总营收'] + safe_float(row.get('总营收')), 2)
            suyang_this_for_wording['市场GMV'] = round(suyang_this_for_wording['市场GMV'] + safe_float(row.get('市场GMV')), 2)
            suyang_this_for_wording['辅导GMV'] = round(suyang_this_for_wording['辅导GMV'] + safe_float(row.get('辅导GMV')), 2)
    max_mkt = 0
    max_mkt_name = ""
    suyang_mkt_total = suyang_this_for_wording['市场GMV'] or 0
    for sname in suyang:
        for row in this_team:
            if sname in str(row.get("项目组", "")):
                mkt = safe_float(row.get("市场GMV"))
                if mkt > max_mkt:
                    max_mkt = mkt
                    max_mkt_name = sname
                break
    定性 = "[待分析]"

    # ===== 整体summary line =====
    lines.append(f"{suyang_short}本周整体总营收{fmt_wan(suyang_this_for_wording['总营收'])}万（环比{emoji(chg_s_rev)}{fmt_pct(chg_s_rev)}），市场GMV {fmt_wan(suyang_this_for_wording['市场GMV'])}万（环比{emoji(chg_s_mkt)}{fmt_pct(chg_s_mkt)}），辅导GMV {fmt_wan(suyang_this_for_wording['辅导GMV'])}万（环比{emoji(chg_s_ment)}{fmt_pct(chg_s_ment)}），{定性}。")
    lines.append("")

    # 素养项目组市场GMV明细（全部7个，按固定顺序，直接显示Excel原始值）
    for sname in suyang:
        for row in this_team:
            if sname in str(row.get("项目组", "")):
                mkt = safe_float(row.get("市场GMV"))
                smroi_v = row.get("市场SMROI", "-")
                smroi_num = safe_float(smroi_v) if str(smroi_v) not in ("", "-", "None") else None
                smroi_str = fmt_smroi(smroi_num) if smroi_num is not None else "-"
                smroi_chg_str = row.get("市场SMROI_环比", "-")
                lines.append(f"* {sname}市场GMV {fmt_wan(mkt)}万，SMROI {smroi_str}（{smroi_chg_str}）")
                break

    lines.append("```")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ============================================================
    # 核心洞察与建议
    # ============================================================
    lines.append("## 核心洞察与建议")
    lines.append("")
    # Auto-fill 核心洞察与建议（全动态计算）
    shengxuan_rev2 = sum(safe_float(r.get('总营收', 0)) for r in this_fin if r.get('项目组') in ['全国大班','云领学','强基业务','研习所','新世界'])
    shengxuan_share2 = shengxuan_rev2 / tw_total * 100 if tw_total else 0
    dqb_share2 = 0.0
    for r in this_fin:
        if '全国大班' in str(r.get('项目组', '')):
            dqb_share2 = safe_float(r.get('总营收', 0)) / tw_total * 100 if tw_total else 0
            break

    # ===== 动态计算Q1续报结算尾款 =====
    if has_last_week and last_fin_total:
        q1_tail = safe_float(last_fin_total.get("续报GMV", 0))
        q1_tail_str = f"约{fmt_wan(q1_tail)}万"
    else:
        q1_tail_str = "数据暂不可比"

    # ===== 动态获取升学SMROI同比 =====
    yoy_shengxue_smroi_chg = "数据暂不可比"
    yoy_suyang_chg = "数据暂不可比"
    if yoy_rows:
        sx_yoy_row = get_yoy_biz(yoy_rows, shengxue_short)
        if sx_yoy_row:
            smroi_26 = safe_float(sx_yoy_row.get('_col_9', 0))
            smroi_25 = safe_float(sx_yoy_row.get('_col_10', 0))
            if smroi_25 != 0:
                yoy_shengxue_smroi_chg = f"{smroi_26 - smroi_25:+.2f}"
        sy_yoy_row = get_yoy_biz(yoy_rows, suyang_short)
        if sy_yoy_row:
            yoy_suyang_chg = sy_yoy_row.get('_col_5', '数据暂不可比')

    # ===== 动态计算高优先级：找最严重风险项目 =====
    risk_items = []
    for r in this_fin:
        name = str(r.get('项目组', ''))
        if not name or name in ('', '合计', '升学', '素养'):
            continue
        rev = safe_float(r.get('总营收', 0))
        smroi_v = safe_float(r.get('SMROI', 0)) if str(r.get('SMROI', '-')) not in ('', '-', 'None') else None
        if rev < 0:
            risk_items.append(('GMV为负', name, rev))
        elif smroi_v is not None and smroi_v < 1.0:
            risk_items.append(('SMROI亏损', name, smroi_v))
    risk_items.sort(key=lambda x: (0 if x[0]=='GMV为负' else 1, abs(x[2])), reverse=True)
    if risk_items:
        risk_type, risk_name, risk_val = risk_items[0]
        if risk_type == 'GMV为负':
            high_priority = f"{risk_name}GMV{risk_val:.1f}万为负，需排查原因"
        else:
            high_priority = f"{risk_name}SMROI {risk_val:.2f}<1，投放效率亏损，需确认投放策略"
    else:
        high_priority = "各项目组整体无严重风险，持续观察即可"

    # ===== 动态计算持续关注：找SMROI最高的非负项目 =====
    best_smroi = None
    best_name = None
    for r in this_fin:
        name = str(r.get('项目组', ''))
        if not name or name in ('', '合计', '升学', '素养'):
            continue
        smroi_v = safe_float(r.get('SMROI', 0)) if str(r.get('SMROI', '-')) not in ('', '-', 'None') else None
        rev = safe_float(r.get('总营收', 0))
        if smroi_v is not None and smroi_v > 0 and (best_smroi is None or smroi_v > best_smroi):
            best_smroi = smroi_v
            best_name = name
    if best_name and best_smroi:
        ongoing_note = f"{best_name}（SMROI {best_smroi:.2f}）效率最优，值得关注"
    else:
        ongoing_note = "各项目组SMROI均无显著亮点"

    # ===== 动态计算下周重点观察 =====
    watch_items = []
    if has_last_week and last_fin_total:
        lw_xubao = safe_float(last_fin_total.get('续报GMV', 0))
        if lw_xubao > 0 and tw_xubao < lw_xubao * 0.5:
            watch_items.append(f"续报从{lw_xubao:.0f}万降至{tw_xubao:.0f}万，萎缩超50%")
    if tw_laixin < 1500:
        watch_items.append(f"拉新{tw_laixin:.0f}万偏低，低于1500万警戒线")
    if not watch_items:
        watch_note = "下周重点观察拉新是否回升、续报萎缩是否符合季度结转规律"
    else:
        watch_note = "；".join(watch_items)

    # ===== 动态判断信号 =====
    signal_positive = safe_float(yoy_shengxue_smroi_chg) > 0 if yoy_shengxue_smroi_chg not in ('数据暂不可比', '') else False
    signal_label = "积极" if signal_positive else "需关注"

    lines.append("**Q{quarter}首周核心判断：[待分析]**")
    lines.append("**优先级排序：**")
    lines.append("🔴 **高优先级：[待分析]**")
    lines.append("🟡 **中优先级：[待分析]**")
    lines.append("🟢 **持续关注：[待分析]**")
    lines.append("**下周重点观察：[待分析]**")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append(f"**报告生成时间：{today}**")
    lines.append(f"**数据来源：{data.get('file', '未知')}**")
    lines.append("**分析工具：经营分析Skill（纯模板生成，模型无关）**")
    lines.append("**包含模块：时间进度基准、单周营收趋势、素养项目组分析、GMV进度及SMROI达标、复利跟踪、同比分析、课程业务分析、周报模板**")
    lines.append("")
    lines.append("> 📌 **模块4提示**：若需补充复利跟踪数据，请提供上周的 `项目组画像卡.md` 文件，报告将自动更新模块4。")

    return "\n".join(lines)


# ============================================================
# 主函数
# ============================================================

def resolve_snapshot_path(arg):
    """解析输入参数为 data_snapshot.md 路径。
    
    - 如果参数是 .md 文件 → 直接使用
    - 如果参数是 .xlsx 文件 → 报错提示先运行 read_excel.py
    - 如果参数是目录 → 自动补为 <dir>/data_snapshot.md（兼容旧用法）
    """
    if arg.endswith('.xlsx') or arg.endswith('.xls'):
        print("❌ generate_report.py 只接受 data_snapshot.md，请先运行 read_excel.py", file=sys.stderr)
        sys.exit(1)
    
    if os.path.isdir(arg):
        return os.path.join(arg, "data_snapshot.md"), arg
    
    if os.path.isfile(arg) and arg.endswith('.md'):
        return arg, os.path.dirname(os.path.abspath(arg))
    
    # Fallback: treat as directory
    if os.path.isdir(arg):
        return os.path.join(arg, "data_snapshot.md"), arg
    
    return arg, os.path.dirname(os.path.abspath(arg))


def main():
    if len(sys.argv) < 2:
        print("用法: python generate_report.py <data_snapshot.md> [--output <输出文件路径>]")
        print("  兼容: python generate_report.py <工作目录> [--output <输出文件路径>]")
        sys.exit(1)

    snapshot_path, work_dir = resolve_snapshot_path(sys.argv[1])
    output_path = None

    if '--output' in sys.argv:
        idx = sys.argv.index('--output')
        if idx + 1 < len(sys.argv):
            output_path = sys.argv[idx + 1]

    if not os.path.exists(snapshot_path):
        print(f"❌ 数据快照不存在: {snapshot_path}", file=sys.stderr)
        print("   请先运行 read_excel.py 生成 data_snapshot.md", file=sys.stderr)
        sys.exit(1)

    print(f"📂 数据快照: {snapshot_path}", file=sys.stderr)
    print(f"ℹ️ 当前处理真实数据；如需脱敏请先运行 desensitize.py", file=sys.stderr)

    data = load_from_snapshot(snapshot_path)
    if data is None:
        print(f"❌ data_snapshot.md 中未找到 DATA_JSON 数据块", file=sys.stderr)
        sys.exit(1)
    data["_work_dir"] = work_dir
    cutoff_date = data.get("cutoff_date", "")

    print(f"📅 数据截止日期: {cutoff_date}", file=sys.stderr)

    report = generate_report(data)

    if not output_path:
        week_label, _ = extract_week_number(cutoff_date)
        output_path = os.path.join(work_dir, f"{week_label}周报分析_{cutoff_date}.md")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(report)

    print(f"📄 报告已生成: {output_path}", file=sys.stderr)
    print(f"   共 {len(report.split(chr(10)))} 行", file=sys.stderr)

    # 生成抽检清单
    checklist_path = output_path.replace(".md", "_抽检清单.md")
    generate_spot_check(data, checklist_path)
    print(f"✅ 抽检清单: {checklist_path}", file=sys.stderr)

    # 生成模块6的Excel版本（方便粘贴石墨文档）
    excel_path = output_path.replace(".md", "_模块6.xlsx")
    generate_module6_excel(data, excel_path)
    print(f"✅ 模块6 Excel: {excel_path}", file=sys.stderr)

    # 自动校验循环 (babysit-report)
    auto_validate_and_fix(output_path, snapshot_path)

    # 追加分析日志
    append_analysis_log(output_path, data)


def generate_spot_check(data, output_path):
    """生成关键数字抽检清单，方便人工30秒核对"""

    quarter = data.get("time_progress", {}).get("quarter", 1)
    fin = {r.get("项目组", ""): r for r in data["data"]["周营收-财务口径"]["rows"]}
    fin_lw = {}
    lwd = data.get("last_week_data", {})
    if lwd and "周营收-财务口径" in lwd:
        fin_lw = {r.get("项目组", ""): r for r in lwd["周营收-财务口径"].get("rows", [])}

    qb = {r.get("季度数据", ""): r for r in data["data"].get(f"Q{quarter}季度-B标", {}).get("rows", [])}

    def sf(v):
        try: return float(v)
        except: return 0

    c = fin.get("合计", {})
    c_lw = fin_lw.get("合计", {})
    qb_total = qb.get("总计", {})

    rev = sf(c.get("总营收", 0))
    rev_lw = sf(c_lw.get("总营收", 0))
    xin = sf(c.get("拉新GMV", 0))
    smroi = sf(c.get("SMROI", 0))
    q_gmv = sf(qb_total.get("总GMV", 0))
    q_target = sf(qb_total.get("目标", 0))
    q_achieve = safe_div_pct(q_gmv, q_target)

    # Find top1 change
    top1_name, top1_change = "", 0
    for name in fin:
        if name in ("合计", "") or str(name).startswith("undefined"):
            continue
        x = sf(fin[name].get("拉新GMV", 0))
        x_lw = sf(fin_lw.get(name, {}).get("拉新GMV", 0))
        chg = x - x_lw
        if abs(chg) > abs(top1_change):
            top1_name, top1_change = name, chg

    lines = [
        "# 📋 关键数字抽检清单",
        "",
        "> 请花30秒核对以下5个数字。这5个对了，其他基本不会错。",
        "",
        "| # | 指标 | 报告中的值 | 核对方法 |",
        "|---|------|----------|---------|",
        f"| 1 | 总Revenue | {rev:,.1f}万 | Excel 周营收-财务口径 → 合计行 → 总营收列 |",
        f"| 2 | 拉新GMV | {xin:,.1f}万 | Excel 周营收-财务口径 → 合计行 → 拉新GMV列 |",
        f"| 3 | SMROI | {smroi:.2f} | Excel 周营收-财务口径 → 合计行 → SMROI列 |",
        f"| 4 | 拉新Top1变动 | {top1_name} {top1_change:+,.1f}万 | 本周拉新GMV - 上周拉新GMV |",
        f"| 5 | Q{quarter}达成率 | {q_achieve} | Excel Q{quarter}季度-B标 → 总计行 → 进度%列 |",
        "",
        "### 核对结果",
        "- [ ] 1. 总Revenue ✅/❌",
        "- [ ] 2. 拉新GMV ✅/❌",
        "- [ ] 3. SMROI ✅/❌",
        "- [ ] 4. 拉新Top1变动 ✅/❌",
        f"- [ ] 5. Q{quarter}达成率 ✅/❌",
        "",
        "**全部打✓ → 报告可用 | 任一❌ → 告诉我哪个不对，我修**",
    ]

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def generate_module6_excel(data, output_path):
    """生成模块6（课程业务分析）的Excel版本，方便粘贴到石墨文档"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # 动态季度检测
    quarter = data.get("time_progress", {}).get("quarter", 1)

    wb = openpyxl.Workbook()

    # === Sheet 1: 本周速览 ===
    ws1 = wb.active
    ws1.title = "本周速览"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a73e8", fill_type="solid")
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Get data
    fin = {r.get("项目组", ""): r for r in data["data"]["周营收-财务口径"]["rows"]}
    team = {r.get("项目组", ""): r for r in data["data"]["周营收-团队口径"]["rows"]}
    fin_lw = {}
    team_lw = {}
    lwd_excel = data.get("last_week_data", {})
    if lwd_excel:
        if "周营收-财务口径" in lwd_excel:
            fin_lw = {r.get("项目组", ""): r for r in lwd_excel["周营收-财务口径"].get("rows", [])}
        if "周营收-团队口径" in lwd_excel:
            team_lw = {r.get("项目组", ""): r for r in lwd_excel["周营收-团队口径"].get("rows", [])}

    qb = {r.get("季度数据", ""): r for r in data["data"].get(f"Q{quarter}季度-B标", {}).get("rows", [])}

    c = fin.get("合计", {})
    c_lw = fin_lw.get("合计", {})
    tc = team.get("合计", {})
    tc_lw = team_lw.get("合计", {})
    qb_total = qb.get("总计", {})

    def sf(v):
        try: return float(v)
        except: return 0

    def pct_str(new, old):
        if not old or sf(old) == 0: return "-"
        return f"{(sf(new)-sf(old))/abs(sf(old))*100:+.1f}%"

    # Headers
    headers = ["指标", "本周实际", "上周实际", "环比", f"Q{quarter}累计", f"Q{quarter}预算", "达成率"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data rows
    rev = sf(c.get("总营收", 0))
    rev_lw = sf(c_lw.get("总营收", 0))
    xin = sf(c.get("拉新GMV", 0))
    xin_lw = sf(c_lw.get("拉新GMV", 0))
    xu = sf(c.get("续报GMV", 0))
    xu_lw = sf(c_lw.get("续报GMV", 0))
    smroi_team = sf(tc.get("市场SMROI", 0))
    smroi_team_lw = sf(tc_lw.get("市场SMROI", 0))
    smroi_proj = sf(c.get("SMROI", 0))
    smroi_proj_lw = sf(c_lw.get("SMROI", 0))
    q_gmv = sf(qb_total.get("总GMV", 0))
    q_target = sf(qb_total.get("目标", 0))
    q_achieve = safe_div_pct(q_gmv, q_target)

    # 从Q{quarter}累计同比读取SMROI Q{quarter}累计值（避免硬编码）
    q_smroi_mkt = qb_total.get("拉新SMROI", "1.33")  # 市场SMROI Q{quarter}累计
    q_smroi_target = qb_total.get("SMROI目标", "1.37")  # SMROI目标
    if isinstance(q_smroi_mkt, str) and q_smroi_mkt.replace(".", "").replace("-", "").isdigit():
        q_smroi_mkt_fmt = f"{float(q_smroi_mkt):.2f}"
    else:
        q_smroi_mkt_fmt = str(q_smroi_mkt)
    if isinstance(q_smroi_target, str) and q_smroi_target.replace(".", "").replace("-", "").isdigit():
        q_smroi_target_fmt = f"{float(q_smroi_target):.2f}"
    else:
        q_smroi_target_fmt = str(q_smroi_target)

    # 从Q{quarter}季度-B标总计行读取拉新/续报Q数据
    q_laixin_qtd = sf(qb_total.get("拉新QTD", 0))
    q_laixin_budget = sf(qb_total.get("预算", 0))
    q_laixin_progress = qb_total.get("拉新进度", "-")
    q_xubao_qtd = sf(qb_total.get("续费QTD", 0))
    q_xubao_budget = sf(qb_total.get("续费QTD_预算", 0))
    q_xubao_progress = qb_total.get("续报进度", "-")

    # Bug 3 Fix: ROI/SMROI rows use subtraction (actual - target) for 达成率, not division
    # GMV rows still use percentage
    q_smroi_mkt_val = sf(q_smroi_mkt) if q_smroi_mkt not in ('-', '', None) else None
    q_smroi_target_val = sf(q_smroi_target) if q_smroi_target not in ('-', '', None) else None
    if q_smroi_mkt_val is not None and q_smroi_target_val is not None and q_smroi_target_val != 0:
        smroi_achieve_excel = q_smroi_mkt_val - q_smroi_target_val  # subtraction, not division
    else:
        smroi_achieve_excel = None

    # Helper: decide if a value should be written as numeric (float) or kept as string
    # Returns (display_value, is_numeric, numeric_val)
    def cell_val(v, decimals=1, is_pct=False):
        """Convert to proper cell value. v can be numeric or string '-'."""
        if v == "-" or v == "" or v is None:
            return ("-", False, None)
        try:
            fv = float(v)
            return (round(fv, decimals), True, fv)
        except (ValueError, TypeError):
            return (str(v), False, None)

    rows_data = [
        # label, 本周实际(numeric), 上周实际(numeric/str), 环比(str),
        # Q{quarter}累计(numeric), Q{quarter}预算(numeric), 达成率(str/numeric)
        ["GMV（万）",  rev,       rev_lw if rev_lw else "-", pct_str(rev, rev_lw),
         q_gmv, q_target, q_achieve],
        ["其中：拉新", xin,       xin_lw if xin_lw else "-", pct_str(xin, xin_lw),
         q_laixin_qtd if q_laixin_qtd else "-", q_laixin_budget if q_laixin_budget else "-",
         q_laixin_progress if q_laixin_progress != "-" else "-"],
        ["其中：续报", xu,        xu_lw if xu_lw else "-",  pct_str(xu, xu_lw),
         q_xubao_qtd if q_xubao_qtd else "-", q_xubao_budget if q_xubao_budget else "-",
         q_xubao_progress if q_xubao_progress != "-" else "-"],
        # SMROI rows: 达成率 = actual - target (subtraction), not percentage
        ["市场SMROI", round(smroi_team, 2) if smroi_team else "-",
         round(smroi_team_lw, 2) if smroi_team_lw else "-",
         pct_str(smroi_team, smroi_team_lw),
         round(q_smroi_mkt_val, 2) if q_smroi_mkt_val is not None else "-",
         round(q_smroi_target_val, 2) if q_smroi_target_val is not None else "-",
         round(smroi_achieve_excel, 2) if smroi_achieve_excel is not None else "-"],
        ["项目SMROI", round(smroi_proj, 2) if smroi_proj else "-",
         round(smroi_proj_lw, 2) if smroi_proj_lw else "-",
         pct_str(smroi_proj, smroi_proj_lw),
         round(q_smroi_mkt_val, 2) if q_smroi_mkt_val is not None else "-",
         round(q_smroi_target_val, 2) if q_smroi_target_val is not None else "-",
         round(smroi_achieve_excel, 2) if smroi_achieve_excel is not None else "-"],
    ]

    # Bug Fix (v6.4): SMROI rows must use decimals=2 in cell_val to preserve precision.
    # With decimals=1, smroi_achieve_excel=-0.02 would round to -0.0 and display as 0.00 in Excel.
    smroi_row_indices = {5, 6}  # Excel row indices for SMROI rows (enumerate starts at 2: GMV=2,拉新=3,续报=4,市场SMROI=5,项目SMROI=6)
    for r_idx, row_data in enumerate(rows_data, 2):
        for c_idx, val in enumerate(row_data, 1):
            # Determine display value and whether it's a number
            # SMROI rows (市场SMROI/项目SMROI) need decimals=2 for the 达成率 column
            dc = 2 if r_idx in smroi_row_indices else 1
            display, is_num, num_val = cell_val(val, decimals=dc)
            cell = ws1.cell(row=r_idx, column=c_idx, value=display if not is_num else display)
            # Apply number_format for numeric cells with value >= 100 (千分位效果)
            if is_num and num_val is not None and abs(num_val) >= 100:
                cell.number_format = '#,##0.0'
            cell.font = bold_font if c_idx == 1 else normal_font
            cell.alignment = left if c_idx == 1 else center
            cell.border = thin_border

    # Auto width
    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = max(max_len * 1.5 + 2, 12)

    # === Sheet 2: 分产品线 ===
    ws2 = wb.create_sheet("分产品线")

    # 动态提取业务线分组（兼容脱敏/原始名称）
    biz_groups = extract_business_groups(data, quarter)
    shengxue = biz_groups["shengxue"]
    suyang = biz_groups["suyang"]
    shengxue_label = biz_groups["shengxue_label"]
    suyang_label = biz_groups["suyang_label"]
    shengxue_short = shengxue_label.split("(")[0].split("（")[0].strip() if shengxue_label else "升学"
    suyang_short = suyang_label.split("(")[0].split("（")[0].strip() if suyang_label else "素养"

    headers2 = ["产品线", "本周GMV（万）", "周环比", f"Q{quarter}累计（万）", f"Q{quarter}预算达成率", "上年同期（万）", "季度同比", "备注"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Read YoY from data_snapshot JSON (same desensitized names as other data)
    yoy_data = {}
    yoy_sheet = data.get("data", {}).get(f"Q{quarter}累计同比", {})
    yoy_rows = yoy_sheet.get("rows", [])
    for row in yoy_rows:
        # First column key is "季度同比数据" (the header name from read_excel)
        name_raw = row.get("季度同比数据", "")
        if not name_raw:
            continue
        name = str(name_raw).strip().replace("  └ ", "").replace("└ ", "")
        # _col_3 = 2026 GMV, _col_4 = 2025 GMV, _col_5 = 同比增长
        gmv_26 = row.get("_col_3")
        gmv_25 = row.get("_col_4")
        change = row.get("_col_5", "-")
        if gmv_26 is not None and gmv_26 != "" and name not in ("", "指标", "总GMV", "拉新GMV", "续报GMV", "平均SMROI", "业务线/项目组", "LCT整体"):
            yoy_data[name] = {
                "this_gmv": sf(gmv_26),
                "last_gmv": sf(gmv_25),
                "change": change if change else "-"
            }
    if yoy_data:
        print(f"  ✅ YoY数据已加载: {len(yoy_data)}个项目组", file=sys.stderr)
    else:
        print(f"  ⚠️ YoY数据为空，Q{quarter}累计同比sheet可能缺失", file=sys.stderr)

    row_idx = 2

    def write_product_row(ws, r, name, is_bold=False, indent=""):
        f_row = fin.get(name, {})
        f_lw_row = fin_lw.get(name, {})
        this_rev = sf(f_row.get("总营收", 0))
        mom = f_row.get("环比", "-")

        qb_row = qb.get(name, {})
        if not qb_row:
            for k, v in qb.items():
                if name in str(k):
                    qb_row = v
                    break
        qb_gmv_v = sf(qb_row.get("总GMV", 0))
        qb_rate = qb_row.get("进度%", "-")

        yoy = yoy_data.get(name, {})
        last_gmv = yoy.get("last_gmv", 0) or 0
        yoy_chg = yoy.get("change", "-")

        display_name = f"{indent}{name}"
        # Numeric columns: 2=本周GMV, 4=Q{quarter}累计, 6=上年同期
        vals = [display_name, this_rev, str(mom), qb_gmv_v, str(qb_rate), last_gmv if last_gmv else "-", str(yoy_chg), ""]
        numeric_cols = {2, 4, 6}  # 1-indexed

        font_to_use = bold_font if is_bold else normal_font
        for c, val in enumerate(vals, 1):
            is_num = c in numeric_cols and val != "-"
            display = val if is_num else str(val) if not is_num else val
            cell = ws.cell(row=r, column=c, value=display if not isinstance(display, (int, float)) else display)
            # Bug 1 Fix: apply #,##0.0 number_format for numeric cells >= 100
            if is_num and isinstance(display, (int, float)) and abs(display) >= 100:
                cell.number_format = '#,##0.0'
            cell.font = font_to_use
            cell.alignment = left if c == 1 else center
            cell.border = thin_border
        return r + 1

    # Shengxue subtotal
    sx_rev = sum(sf(fin.get(n, {}).get("总营收", 0)) for n in shengxue)
    sx_rev_lw = sum(sf(fin_lw.get(n, {}).get("总营收", 0)) for n in shengxue)
    sx_mom = f"{(sx_rev-sx_rev_lw)/abs(sx_rev_lw)*100:+.1f}%" if sx_rev_lw else "-"
    sx_q1 = qb.get(shengxue_label, {})
    sx_yoy = yoy_data.get(shengxue_short, {})

    sx_last_gmv = sx_yoy.get("last_gmv", 0) or 0
    vals = [shengxue_short, sx_rev, sx_mom,
            sf(sx_q1.get('总GMV', 0)), str(sx_q1.get("进度%", "-")),
            sx_last_gmv if sx_last_gmv else "-",
            str(sx_yoy.get("change", "-")), ""]
    # Numeric cols in Sheet 2: 2, 4, 6
    for c, val in enumerate(vals, 1):
        is_num = c in {2, 4, 6} and val != "-"
        cell = ws2.cell(row=row_idx, column=c, value=val if is_num else str(val) if not isinstance(val, (int, float)) else val)
        if is_num and isinstance(val, (int, float)) and abs(val) >= 100:
            cell.number_format = '#,##0.0'
        cell.font = bold_font
        cell.alignment = left if c == 1 else center
        cell.border = thin_border
    row_idx += 1

    for name in shengxue:
        row_idx = write_product_row(ws2, row_idx, name, indent="  └ ")

    # Suyang subtotal
    sy_rev = sum(sf(fin.get(n, {}).get("总营收", 0)) for n in suyang)
    sy_rev_lw = sum(sf(fin_lw.get(n, {}).get("总营收", 0)) for n in suyang)
    sy_mom = f"{(sy_rev-sy_rev_lw)/abs(sy_rev_lw)*100:+.1f}%" if sy_rev_lw else "-"
    sy_q1 = qb.get(suyang_label, {})
    sy_yoy = yoy_data.get(suyang_short, {})

    sy_last_gmv = sy_yoy.get("last_gmv", 0) or 0
    vals = [suyang_short, sy_rev, sy_mom,
            sf(sy_q1.get('总GMV', 0)), str(sy_q1.get("进度%", "-")),
            sy_last_gmv if sy_last_gmv else "-",
            str(sy_yoy.get("change", "-")), ""]
    for c, val in enumerate(vals, 1):
        is_num = c in {2, 4, 6} and val != "-"
        cell = ws2.cell(row=row_idx, column=c, value=val if is_num else str(val) if not isinstance(val, (int, float)) else val)
        if is_num and isinstance(val, (int, float)) and abs(val) >= 100:
            cell.number_format = '#,##0.0'
        cell.font = bold_font
        cell.alignment = left if c == 1 else center
        cell.border = thin_border
    row_idx += 1

    for name in suyang:
        row_idx = write_product_row(ws2, row_idx, name, indent="  └ ")

    # Auto width
    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = max(max_len * 1.5 + 2, 12)

    # Sanity check: warn if all product line GMV values are zero
    total_gmv = sx_rev + sy_rev
    if total_gmv == 0 and len(shengxue) + len(suyang) > 0:
        print("  ⚠️ 模块6分产品线GMV全为0！可能是数据匹配问题", file=sys.stderr)
        print(f"     fin keys: {list(fin.keys())}", file=sys.stderr)
        print(f"     shengxue: {shengxue}", file=sys.stderr)
        print(f"     suyang: {suyang}", file=sys.stderr)

    wb.save(output_path)

    # P1增强：生成后基本校验（存在/非空/必要工作表/关键单元格）
    import os as _os
    if not _os.path.exists(output_path):
        raise RuntimeError(f"❌ 模块6 Excel生成失败：文件不存在 {output_path}")
    if _os.path.getsize(output_path) < 1024:
        raise RuntimeError(f"❌ 模块6 Excel生成失败：文件小于1KB（疑似空文件） {output_path}")

    # 重新打开验证内容完整性
    wb_check = openpyxl.load_workbook(output_path)
    required_sheets = ["本周速览", "分产品线"]
    for rs in required_sheets:
        if rs not in wb_check.sheetnames:
            raise RuntimeError(f"❌ 模块6 Excel缺少必要工作表：{rs}")
        ws_check = wb_check[rs]
        # 至少应有表头行(1) + 数据行(2)
        if ws_check.max_row < 2 or ws_check.max_column < 2:
            raise RuntimeError(f"❌ 模块6 Excel工作表 {rs} 数据不足（{ws_check.max_row}行 x {ws_check.max_column}列）")
        # 检查关键单元格非空（本周速览B2=总营收，Q1累计F2=预算）
        if rs == "本周速览":
            b2 = ws_check.cell(row=2, column=2).value
            if b2 is None or str(b2).strip() in ("", "-"):
                raise RuntimeError(f"❌ 模块6 Excel 本周速览 B2（总营收）为空，内容：{b2}")
    wb_check.close()
    print(f"  ✅ 模块6 Excel 生成校验通过: {output_path}", file=sys.stderr)


# ============================================================
# 报告保姆机制 (babysit-report)
# ============================================================

def auto_validate_and_fix(report_path, snapshot_path, max_rounds=3):
    """自动校验循环：生成报告后自动跑validate_report.py，失败则修复重试"""
    validate_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "validate_report.py")
    if not os.path.exists(validate_script):
        print("⏭️  validate_report.py 不存在，跳过自动校验", file=sys.stderr)
        return True

    for round_num in range(1, max_rounds + 1):
        print(f"\n🔄 校验轮次 {round_num}/{max_rounds}...", file=sys.stderr)
        result = subprocess.run(
            [sys.executable, validate_script, report_path, snapshot_path],
            capture_output=True, text=True, timeout=60
        )
        if result.stdout:
            print(result.stdout, file=sys.stderr)
        if result.stderr:
            print(result.stderr, file=sys.stderr)

        if result.returncode == 0:
            print(f"✅ 校验通过（第{round_num}轮）", file=sys.stderr)
            return True
        else:
            print(f"⚠️  校验未通过（第{round_num}轮），returncode={result.returncode}", file=sys.stderr)
            if round_num < max_rounds:
                print("   尝试重新生成报告...", file=sys.stderr)
                try:
                    data = load_from_snapshot(snapshot_path)
                    if data:
                        data["_work_dir"] = os.path.dirname(os.path.abspath(snapshot_path))
                        report_text = generate_report(data)
                        with open(report_path, "w", encoding="utf-8") as f:
                            f.write(report_text)
                        print("   已重新生成报告", file=sys.stderr)
                    else:
                        print("   无法从快照加载数据", file=sys.stderr)
                        break
                except Exception as e:
                    print(f"   重新生成失败: {e}", file=sys.stderr)
                    break

    print(f"❌ 校验未通过，已达最大轮次 {max_rounds}", file=sys.stderr)
    return False


# ============================================================
# 分析日志 (analysis_log)
# ============================================================

def append_analysis_log(report_path, data):
    """追加分析日志到 analysis_log.jsonl"""
    log_path = os.path.join(_log_dir(), "analysis_log.jsonl")
    os.makedirs(os.path.dirname(log_path), exist_ok=True)

    tp = data.get("time_progress", {})
    q_progress = tp.get("q_progress", 0)
    quarter = tp.get("quarter", 1)

    # 从数据中提取关键指标
    fin_total = get_total_row(data, "周营收-财务口径")
    tw_total = safe_float(fin_total.get("总营收") if fin_total else 0)
    tw_smroi = safe_float(fin_total.get("SMROI") if fin_total else 0)

    # 风险扫描
    risks = scan_risks(data, q_progress, tp.get("month_progress", 0), quarter)
    key_risk = risks[0] if risks else "无明显风险"

    entry = {
        "skill": "jingying-fenxi",
        "date": datetime.now().strftime("%Y-%m-%d"),
        "business": "课程",
        "week_revenue": round(tw_total, 1),
        "smroi": round(tw_smroi, 2),
        f"Q{quarter}_progress": q_progress,
        "key_risk": key_risk[:50],
        "key_action": "持续监控",
        "report_path": report_path,
    }
    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
        print(f"📝 分析日志已追加: {log_path}", file=sys.stderr)
    except Exception as e:
        print(f"⚠️  分析日志写入失败: {e}", file=sys.stderr)


def get_last_analysis_log():
    """读取上一条本Skill的分析日志"""
    log_path = os.path.join(_log_dir(), "analysis_log.jsonl")
    if not os.path.exists(log_path):
        return None
    try:
        last = None
        with open(log_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                entry = json.loads(line)
                if entry.get("skill") == "jingying-fenxi":
                    last = entry
        return last
    except Exception:
        return None


# ============================================================
# 使用频次追踪 (usage tracking)
# ============================================================

def track_usage(action="generate_report"):
    """追踪Skill调用频次"""
    usage_path = _usage_path()
    try:
        with open(usage_path, "a", encoding="utf-8") as f:
            f.write(json.dumps({
                "skill": "jingying-fenxi",
                "time": datetime.now().isoformat(),
                "action": action
            }, ensure_ascii=False) + "\n")
    except Exception:
        pass


if __name__ == "__main__":
    # Usage tracking
    track_usage()
    main()
